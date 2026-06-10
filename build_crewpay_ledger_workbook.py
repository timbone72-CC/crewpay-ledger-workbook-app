from pathlib import Path
from datetime import date, datetime, time
import sys


LOCAL_DEPS = Path(__file__).resolve().parent / ".deps" / "usr" / "lib" / "python3" / "dist-packages"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "CrewPay_Ledger_Workbook.xlsx"
SHEET_ORDER = [
    "Instructions",
    "Dashboard",
    "Workers",
    "Jobs",
    "Time Entries",
    "Pay Periods",
    "Worker Proof",
    "Access Status Demo",
    "Workflow Demo",
    "Proof Exports",
    "Access Log",
    "Correction Log",
    "Schedule",
    "Admin Notices",
    "Calendar Sync Log",
    "App Config",
    "Pending Worker Intake",
    "Pending Pay Period Intake",
    "Pending Time Entries",
    "App Submission Log",
    "Bridge Schema",
    "Dropdown Lists",
]

DROPDOWNS = {
    "Access Status": ["Active", "Inactive"],
    "Job Status": ["Active", "Paused", "Closed"],
    "Approval Status": ["Draft", "Submitted", "Approved", "Finalized", "Paid", "Rejected"],
    "Pay Period Status": ["Open", "Finalized", "Paid"],
    "Payment Status": ["Unpaid", "Paid"],
    "Export Type": ["Print", "PDF", "CSV"],
    "Schedule Status": ["Planned", "Scheduled", "Completed", "Canceled"],
    "Recipient Type": ["Worker", "All Active Workers"],
    "Delivery Method": ["Workbook", "Email Ready", "Gmail Sent"],
    "Notice Status": ["Draft", "Posted", "Sent", "Archived"],
    "Sync Status": ["Not Synced", "Synced", "Failed", "Skipped"],
    "Submission Status": ["Pending", "Reviewed", "Accepted", "Rejected"],
    "Boolean Flag": ["TRUE", "FALSE"],
}

THEME = {
    "navy": "12324A",
    "navy_2": "1E4E6A",
    "blue": "DCEBF7",
    "blue_2": "EEF6FC",
    "card": "FFFFFF",
    "grid": "CAD6E2",
    "text": "1F2937",
    "muted": "6B7280",
    "green": "D9EAD3",
    "green_text": "166534",
    "yellow": "FFF2CC",
    "orange": "FCE4D6",
    "red": "F8D7DA",
    "gray": "E5E7EB",
    "gray_2": "F8FAFC",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=THEME["grid"]),
    right=Side(style="thin", color=THEME["grid"]),
    top=Side(style="thin", color=THEME["grid"]),
    bottom=Side(style="thin", color=THEME["grid"]),
)

WORKERS = [
    ["W-1001", "Maya Ellis", "maya.ellis@example.local", "Crew Lead", "Active", "2026-06-01", "", "Active worker for current entries."],
    ["W-1002", "Leo Grant", "leo.grant@example.local", "Installer", "Active", "2026-06-01", "", "Active worker with pending approval."],
    ["W-1003", "Priya Shah", "priya.shah@example.local", "Helper", "Inactive", "2026-05-15", "2026-06-09", "Inactive worker with historical proof records."],
]

JOBS = [
    ["J-2001", "Oak Ridge Units", "Oak Ridge Apartments", "Active", 32, "", "Apartment unit prep and repairs."],
    ["J-2002", "Bluebird Cafe Refresh", "Bluebird Cafe", "Active", 30, "", "Cafe refresh punch list."],
    ["J-2003", "Cedar Lane Repairs", "Cedar Lane Rentals", "Closed", 24, "cal-demo-cedar-001", "Closed job retained for proof history."],
]

TIME_ENTRIES = [
    ["T-3001", "W-1001", "", "J-2001", "", "2026-06-02", "08:00", "16:30", 30, "", "", "", 35, 0, "", "Approved", "2026-06-02 16:45", "2026-06-03 09:10", "", "Unit prep and fixture checks."],
    ["T-3002", "W-1002", "", "J-2001", "", "2026-06-02", "08:00", "15:00", 30, "", "", "", 0, 0, "", "Submitted", "2026-06-02 15:20", "", "", "Drywall patch support pending approval."],
    ["T-3003", "W-1003", "", "J-2003", "", "2026-06-03", "09:00", "14:00", 0, "", "", "", 0, 0, "", "Approved", "2026-06-03 14:10", "2026-06-04 08:30", "", "Historical approved work before inactive status."],
    ["T-3004", "W-1001", "", "J-2002", "", "2026-06-06", "07:30", "13:30", 0, "", "", "", 0, 0, "", "Paid", "2026-06-06 13:40", "2026-06-07 10:00", "Paid period proof generated after final punch list.", "Final punch list."],
    ["T-3005", "W-1003", "", "J-2003", "", "2026-06-05", "08:30", "12:30", 0, "", "", "", 15, 0, "", "Finalized", "2026-06-05 12:45", "2026-06-06 09:15", "Visible correction: added reimbursement before worker was inactive.", "Historical reimbursement correction retained."],
]

PAY_PERIODS = [
    ["P-4001", "W-1001", "", "2026-06-01", "2026-06-07", "Paid", "Paid", "", "", "", "", "", "2026-06-07 10:00", "2026-06-08 12:00", "Paid worker proof can be regenerated."],
    ["P-4002", "W-1002", "", "2026-06-01", "2026-06-07", "Open", "Unpaid", "", "", "", "", "", "", "", "Open period with submitted entry."],
    ["P-4003", "W-1003", "", "2026-06-01", "2026-06-07", "Finalized", "Unpaid", "", "", "", "", "", "2026-06-08 09:00", "", "Inactive worker history remains visible and provable."],
]

PROOF_EXPORTS = [
    ["X-5001", "W-1001", "Maya Ellis", "P-4001", "Print", "2026-06-08 12:05", "Admin Demo", "Printed proof packet", "Worker-specific proof generated from workbook records."],
    ["X-5002", "W-1003", "Priya Shah", "P-4003", "CSV", "2026-06-09 09:30", "Admin Demo", "CSV proof export", "Inactive worker proof regenerated after access status changed."],
]

ACCESS_LOG = [
    ["A-6001", "W-1003", "Priya Shah", "Active", "Inactive", "2026-06-09 08:00", "Admin Demo", "End of assignment; historical proof retained."],
]

CORRECTION_LOG = [
    ["C-7001", "T-3005", "W-1003", "Priya Shah", "P-4003", "2026-06-08", "Admin Demo", "Add missing reimbursement", "Reimbursement 0", "Reimbursement 15", "Correction visible in Time Entries and log."],
]

SCHEDULE = [
    ["S-8001", "J-2001", "Oak Ridge Units", "W-1001", "Maya Ellis", "2026-06-11", "08:00", "16:00", "Scheduled", "cal-demo-oak-001", "Planning only; not proof."],
    ["S-8002", "J-2002", "Bluebird Cafe Refresh", "W-1002", "Leo Grant", "2026-06-12", "09:00", "13:00", "Planned", "", "Schedule reference only."],
]

ADMIN_NOTICES = [
    ["N-9001", "2026-06-08 12:15", "Admin Demo", "Worker", "W-1001", "Maya Ellis", "Proof ready", "Your paid proof packet is ready in the workbook.", "P-4001", "Workbook", "Posted", "2026-06-08 12:15", "One-way notice, not chat or proof."],
    ["N-9002", "2026-06-09 08:20", "Admin Demo", "All Active Workers", "", "", "Schedule reminder", "Check the Schedule tab for upcoming planned work.", "", "Email Ready", "Draft", "", "Email-ready text only; no send automation."],
]

CALENDAR_SYNC = [
    ["G-10001", "J-2001", "Oak Ridge Units", "W-1001", "Maya Ellis", "cal-demo-oak-001", "2026-06-11", "Not Synced", "", "Calendar is reference only, not proof."],
]

APP_CONFIG = [
    ["Workbook Source Of Truth", "TRUE", "Workbook remains final authority."],
    ["App Role", "Companion Input Layer", "App submits controlled intake records only."],
    ["Backend Type", "Google Apps Script Web App", "No separate backend server/database."],
    ["Endpoint URL", "PASTE_DEPLOYED_APPS_SCRIPT_WEB_APP_URL_HERE", "Set after Apps Script deployment."],
    ["Bridge Enabled", "FALSE", "Turn TRUE only after deployment/testing."],
    ["Shared Token Required", "TRUE", "Browser-visible token is not full security; use private/demo boundary."],
    ["Intake Review Required", "TRUE", "App submissions should be reviewed before final record use."],
]

PENDING_WORKER_INTAKE = [
    ["PW-0001", "2026-06-10 09:00", "App Demo", "Pending", "W-001", "Sample Worker", "Active", "Crew", "sample@example.com", "Sample pending worker intake row.", "", "", ""],
]

PENDING_PAY_PERIOD_INTAKE = [
    ["PP-0001", "2026-06-10 09:05", "App Demo", "Pending", "PP-001", "W-001", "Sample Worker", "2026-06-01", "2026-06-07", "2026-06-14", "Sample pending pay period row.", "", "", ""],
]

PENDING_TIME_ENTRIES = [
    ["PT-0001", "2026-06-10 09:10", "App Demo", "Pending", "E-001", "W-001", "Sample Worker", "PP-001", "2026-06-03", "Sample Work", 8, 25, "=K2*L2", "Sample pending time entry row.", "", "", ""],
]

APP_SUBMISSION_LOG = [
    ["AL-0001", "2026-06-10 09:15", "healthCheck", "App Demo", "OK", "", "", "", "Sample app submission log row.", "Sample only.", "Not deployed"],
]

BRIDGE_SCHEMA = [
    ["healthCheck", "App Submission Log", "action", "client_version, submission_source", '{ok:true, action:"healthCheck", workbook:"CrewPay Ledger"}', "Bridge disabled; workbook unavailable", "Readiness check only; may log a health check row if desired."],
    ["getWorkbookSchema", "Bridge Schema / Dropdown Lists", "action", "schema_version", "{ok:true, config:{...}, dropdowns:{...}}", "Bridge disabled; schema unavailable", "Return allowed tabs, writable fields, and dropdown values."],
    ["submitWorkerIntake", "Pending Worker Intake", "worker_name, access_status, role_trade, contact", "worker_id, notes, idempotency_key", "{ok:true, intake_id:'PW-####'}", "Missing required field; duplicate submission; invalid access status", "First landing zone for worker records from the app."],
    ["submitPayPeriod", "Pending Pay Period Intake", "pay_period_id, worker_id, period_start, period_end", "worker_name, pay_date, notes, idempotency_key", "{ok:true, intake_id:'PP-####'}", "Invalid date range; missing worker; duplicate submission", "Pending setup only. Final Pay Periods tab remains workbook/admin reviewed."],
    ["submitTimeEntry", "Pending Time Entries", "worker_id, pay_period_id, work_date, job_work_type, hours, rate", "worker_name, entry_id, notes, idempotency_key", "{ok:true, intake_id:'PT-####'}", "Inactive worker; invalid hours/rate; duplicate submission", "Safest first write target for app time entries."],
    ["logProofExport", "Proof Exports", "worker_id, pay_period_id, export_type, export_reference", "worker_name, notes", "{ok:true, export_id:'X-####'}", "Worker/pay period mismatch; invalid export type", "Do not write Worker Proof directly."],
    ["logCorrection", "Correction Log", "entry_id, worker_id, pay_period_id, correction_reason, original_value_summary, new_value_summary", "worker_name, notes", "{ok:true, correction_id:'C-####'}", "Missing proof context; duplicate correction", "Corrections must be visible, not silent."],
    ["logAccessChange", "Access Log", "worker_id, previous_status, new_status, reason", "worker_name", "{ok:true, log_id:'A-####'}", "Invalid status; missing reason", "Future access changes must preserve historical proof."],
    ["createAdminNotice", "Admin Notices", "recipient_type, subject, message", "worker_id, worker_name, related_pay_period_id, notes", "{ok:true, notice_id:'N-####'}", "Invalid recipient type; missing message", "Creates a one-way notice row only; sending is separate."],
]


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_sheet_base(ws, hide_grid=True):
    ws.sheet_view.showGridLines = not hide_grid
    ws.sheet_properties.tabColor = THEME["navy"]
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Aptos", color=THEME["text"])
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def header_style(ws, row=1, start_col=1, end_col=None):
    end_col = end_col or ws.max_column
    fill = PatternFill("solid", fgColor=THEME["navy"])
    font = Font(name="Aptos", bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="medium", color=THEME["navy_2"]))
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def table_style(ws, min_row, max_row, min_col, max_col):
    for row_index, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col), min_row):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index > min_row:
                cell.fill = PatternFill("solid", fgColor=THEME["gray_2"] if row_index % 2 == 0 else THEME["card"])


def title_bar(ws, title, subtitle=None, end_col=8):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["A1"].font = Font(name="Aptos", bold=True, size=20, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    for col in range(1, end_col + 1):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=THEME["navy"])
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws["A2"] = subtitle
        ws["A2"].fill = PatternFill("solid", fgColor=THEME["blue_2"])
        ws["A2"].font = Font(name="Aptos", italic=True, color=THEME["muted"])
        ws.row_dimensions[2].height = 24


def card(ws, row, col, height, width, title, body="", fill=None, title_color=None):
    fill = fill or THEME["card"]
    title_color = title_color or THEME["navy"]
    end_row = row + height - 1
    end_col = col + width - 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    ws.cell(row, col, title)
    ws.cell(row, col).font = Font(name="Aptos", bold=True, size=11, color=title_color)
    ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if height > 1:
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=end_row, end_column=end_col)
        ws.cell(row + 1, col, body)
        ws.cell(row + 1, col).font = Font(name="Aptos", size=10, color=THEME["text"])
        ws.cell(row + 1, col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for r in range(row, end_row + 1):
        for c in range(col, end_col + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, c).border = THIN_BORDER


def metric_card(ws, row, col, label, formula_or_value, width=2, number_format=None, fill=None):
    card(ws, row, col, 2, width, label, "", fill or THEME["card"])
    ws.cell(row + 1, col, formula_or_value)
    ws.cell(row + 1, col).font = Font(name="Aptos", bold=True, size=18, color=THEME["navy"])
    ws.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        ws.cell(row + 1, col).number_format = number_format


def add_status_formatting(ws, cell_range, col_letter):
    rules = [
        ("Active", THEME["green"]),
        ("Approved", THEME["green"]),
        ("Paid", THEME["green"]),
        ("Submitted", THEME["yellow"]),
        ("Draft", THEME["blue"]),
        ("Finalized", THEME["blue"]),
        ("Posted", THEME["blue"]),
        ("Email Ready", THEME["yellow"]),
        ("Scheduled", THEME["yellow"]),
        ("Planned", THEME["blue"]),
        ("Rejected", THEME["red"]),
        ("Inactive", THEME["gray"]),
        ("Archived", THEME["gray"]),
        ("Closed", THEME["gray"]),
        ("Canceled", THEME["gray"]),
        ("Cancelled", THEME["gray"]),
    ]
    for status, color in rules:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'${col_letter}2="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )


def add_validation(ws, list_name, cell_range):
    keys = list(DROPDOWNS.keys())
    col = get_column_letter(keys.index(list_name) + 1)
    end_row = len(DROPDOWNS[list_name]) + 1
    dv = DataValidation(type="list", formula1=f"'Dropdown Lists'!${col}$2:${col}${end_row}", allow_blank=True)
    dv.errorTitle = "Invalid value"
    dv.error = "Choose a value from Dropdown Lists."
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_range_validation(ws, source_range, cell_range):
    dv = DataValidation(type="list", formula1=source_range, allow_blank=True)
    dv.errorTitle = "Invalid value"
    dv.error = "Choose a value from the approved workbook range."
    ws.add_data_validation(dv)
    dv.add(cell_range)


def as_date(value):
    if not value:
        return None
    return date.fromisoformat(str(value))


def as_time(value):
    if not value:
        return None
    hour, minute = str(value).split(":")[:2]
    return time(int(hour), int(minute))


def as_datetime(value):
    if not value:
        return None
    return datetime.strptime(str(value), "%Y-%m-%d %H:%M")


def build_dropdowns(ws):
    ws.title = "Dropdown Lists"
    ws.sheet_view.showGridLines = False
    for col_idx, (header, values) in enumerate(DROPDOWNS.items(), 1):
        ws.cell(1, col_idx, header)
        for row_idx, value in enumerate(values, 2):
            ws.cell(row_idx, col_idx, value)
    header_style(ws)
    table_style(ws, 1, ws.max_row, 1, ws.max_column)
    ws["O1"] = "Demo Config"
    ws["O2"] = "Support/config tab for demo dropdowns. Not part of the public workflow."
    ws["O1"].font = Font(bold=True, color="FFFFFF")
    ws["O1"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["O2"].fill = PatternFill("solid", fgColor=THEME["yellow"])
    ws["O2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["O1"].border = THIN_BORDER
    ws["O2"].border = THIN_BORDER
    set_widths(ws, {get_column_letter(i): 22 for i in range(1, len(DROPDOWNS) + 1)})
    ws.column_dimensions["O"].width = 42
    ws.freeze_panes = "A2"


def build_instructions(ws):
    ws.title = "Instructions"
    title_bar(ws, "CrewPay Ledger", "Public demo guide for a workbook-first worker proof and pay-period ledger.", 10)
    cards = [
        (4, 1, "What this workbook does", "Tracks workers, jobs, time entries, pay periods, proof exports, access changes, correction notes, schedules, and admin notices."),
        (4, 5, "What this workbook does not do", "It is not a full payroll system. It does not include payroll tax logic, HR compliance logic, worker accounts, enterprise permissions, backend storage, or paid APIs."),
        (8, 1, "Source of truth", "The workbook is the Level 1 ledger. Apps and scripts are optional helpers and should write back to workbook records."),
        (8, 5, "Worker proof protection", "Worker Proof is selected by one worker and one pay period. It must never become a crew-wide export."),
        (12, 1, "Inactive worker rule", "Inactive workers lose future-use access only. Their historical time, pay periods, corrections, and proof stay visible."),
        (12, 5, "Correction notes", "Corrections should be visible in Time Entries and Correction Log. No silent overwrite of proof records."),
        (16, 1, "Schedule/calendar note", "Schedule and Google Calendar are planning/reference views only. Time Entries and Pay Periods remain proof."),
        (16, 5, "Admin notices", "Admin Notices are one-way admin-to-worker or admin-to-all notices. They are not chat and not proof."),
        (20, 1, "Level 1 / 1.5 / 2 readiness", "Level 1 is workbook-native. Level 1.5 adds optional Apps Script helpers. Level 2 can later move to a fuller app without changing the proof rules."),
        (20, 5, "Public demo safety note", "All sample names, emails, jobs, notices, and records are fictional demo data. No private worker or client data is included."),
    ]
    for row, col, heading, body in cards:
        card(ws, row, col, 3, 4, heading, body, THEME["card"])
    set_widths(ws, {"A": 20, "B": 16, "C": 16, "D": 16, "E": 20, "F": 16, "G": 16, "H": 16, "I": 12, "J": 12})
    for row in range(1, 24):
        ws.row_dimensions[row].height = 24
    ws.freeze_panes = "A3"


def append_table(ws, headers, rows, widths=None):
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_style(ws, 1, 1, len(headers))
    table_style(ws, 1, max(ws.max_row, 1), 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if widths:
        set_widths(ws, widths)


def build_workers(ws):
    ws.title = "Workers"
    headers = ["Worker ID", "Worker Name", "Worker Email", "Role", "Access Status", "Created At", "Inactive At", "Notes"]
    append_table(ws, headers, WORKERS, {"A": 12, "B": 20, "C": 28, "D": 16, "E": 16, "F": 14, "G": 14, "H": 42})
    add_validation(ws, "Access Status", "E2:E200")
    for row in range(2, ws.max_row + 1):
        ws[f"F{row}"].value = as_date(ws[f"F{row}"].value)
        ws[f"G{row}"].value = as_date(ws[f"G{row}"].value)
        ws[f"F{row}"].number_format = "m/d/yyyy"
        ws[f"G{row}"].number_format = "m/d/yyyy"
    add_status_formatting(ws, "A2:H200", "E")
    card(ws, 1, 10, 3, 3, "Access rule", "Active workers can receive future time entries. Inactive workers are proof-only for historical records.", THEME["blue_2"])


def build_jobs(ws):
    ws.title = "Jobs"
    headers = ["Job ID", "Job Name", "Client or Site", "Status", "Default Rate", "Calendar Event ID", "Notes"]
    append_table(ws, headers, JOBS, {"A": 12, "B": 24, "C": 24, "D": 14, "E": 14, "F": 24, "G": 40})
    add_validation(ws, "Job Status", "D2:D200")
    for row in range(2, ws.max_row + 1):
        ws[f"E{row}"].number_format = "$#,##0.00"
    add_status_formatting(ws, "A2:G200", "D")
    card(ws, 1, 9, 3, 3, "Job rule", "Closed jobs stay visible for proof history. Schedule and calendar references are not proof.", THEME["blue_2"])


def build_time_entries(ws):
    ws.title = "Time Entries"
    headers = ["Entry ID", "Worker ID", "Worker Name", "Job ID", "Job Name", "Work Date", "Start Time", "End Time", "Break Minutes", "Hours", "Rate", "Gross Pay", "Reimbursement", "Deduction", "Net Pay", "Approval Status", "Submitted At", "Approved At", "Correction Note", "Notes"]
    append_table(ws, headers, TIME_ENTRIES, {"A": 12, "B": 12, "C": 20, "D": 12, "E": 24, "F": 13, "G": 12, "H": 12, "I": 14, "J": 12, "K": 12, "L": 14, "M": 15, "N": 12, "O": 14, "P": 16, "Q": 18, "R": 18, "S": 34, "T": 34})
    for row in range(2, 201):
        ws[f"C{row}"] = f'=IFERROR(VLOOKUP(B{row},Workers!$A:$H,2,FALSE),"")'
        ws[f"E{row}"] = f'=IFERROR(VLOOKUP(D{row},Jobs!$A:$G,2,FALSE),"")'
        ws[f"J{row}"] = f'=IF(OR(G{row}="",H{row}=""),"",MAX(0,(H{row}-G{row})*24-(I{row}/60)))'
        ws[f"K{row}"] = f'=IFERROR(VLOOKUP(D{row},Jobs!$A:$G,5,FALSE),"")'
        ws[f"L{row}"] = f'=IF(OR(J{row}="",K{row}=""),"",J{row}*K{row})'
        ws[f"O{row}"] = f'=IF(L{row}="","",L{row}+M{row}-N{row})'
    add_validation(ws, "Approval Status", "P2:P200")
    add_range_validation(ws, "Workers!$A$2:$A$200", "B2:B200")
    add_range_validation(ws, "Jobs!$A$2:$A$200", "D2:D200")
    for row in range(2, ws.max_row + 1):
        ws[f"F{row}"].value = as_date(ws[f"F{row}"].value)
        ws[f"G{row}"].value = as_time(ws[f"G{row}"].value)
        ws[f"H{row}"].value = as_time(ws[f"H{row}"].value)
        ws[f"Q{row}"].value = as_datetime(ws[f"Q{row}"].value)
        ws[f"R{row}"].value = as_datetime(ws[f"R{row}"].value)
    for row in range(2, 201):
        ws[f"F{row}"].number_format = "m/d/yyyy"
        ws[f"G{row}"].number_format = "h:mm"
        ws[f"H{row}"].number_format = "h:mm"
        ws[f"Q{row}"].number_format = "m/d/yyyy h:mm"
        ws[f"R{row}"].number_format = "m/d/yyyy h:mm"
        ws[f"J{row}"].number_format = "0.00"
        for col in ["K", "L", "M", "N", "O"]:
            ws[f"{col}{row}"].number_format = "$#,##0.00"
    ws.conditional_formatting.add("A2:T200", FormulaRule(formula=['$P2="Submitted"'], fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add("A2:T200", FormulaRule(formula=['$P2="Approved"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A2:T200", FormulaRule(formula=['$P2="Paid"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A2:T200", FormulaRule(formula=['$P2="Rejected"'], fill=PatternFill("solid", fgColor=THEME["red"])))
    ws.conditional_formatting.add("A2:T200", FormulaRule(formula=['$P2="Draft"'], fill=PatternFill("solid", fgColor=THEME["blue"])))
    ws.conditional_formatting.add("A2:T200", FormulaRule(formula=['$S2<>""'], fill=PatternFill("solid", fgColor="FCE7E7")))
    ws.conditional_formatting.add(
        "A2:T200",
        FormulaRule(
            formula=['IFERROR(VLOOKUP($B2,Workers!$A:$E,5,FALSE)="Inactive",FALSE)'],
            fill=PatternFill("solid", fgColor="E5E7EB"),
        ),
    )
    card(ws, 1, 22, 4, 4, "Workflow status guide", "Draft = not submitted\nSubmitted = needs admin review\nApproved/Paid = proof-ready\nRejected/corrections stay visible", THEME["blue_2"])


def build_pay_periods(ws):
    ws.title = "Pay Periods"
    headers = ["Pay Period ID", "Worker ID", "Worker Name", "Period Start", "Period End", "Status", "Payment Status", "Total Hours", "Gross Pay", "Reimbursement Total", "Deduction Total", "Net Pay", "Finalized At", "Paid At", "Notes"]
    append_table(ws, headers, PAY_PERIODS, {"A": 15, "B": 12, "C": 20, "D": 14, "E": 14, "F": 14, "G": 14, "H": 13, "I": 14, "J": 20, "K": 18, "L": 14, "M": 18, "N": 18, "O": 42})
    for row in range(2, 101):
        ws[f"C{row}"] = f'=IFERROR(VLOOKUP(B{row},Workers!$A:$H,2,FALSE),"")'
        ws[f"H{row}"] = f'=SUMIFS(\'Time Entries\'!$J:$J,\'Time Entries\'!$B:$B,$B{row},\'Time Entries\'!$F:$F,">="&$D{row},\'Time Entries\'!$F:$F,"<="&$E{row},\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected")'
        ws[f"I{row}"] = f'=SUMIFS(\'Time Entries\'!$L:$L,\'Time Entries\'!$B:$B,$B{row},\'Time Entries\'!$F:$F,">="&$D{row},\'Time Entries\'!$F:$F,"<="&$E{row},\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected")'
        ws[f"J{row}"] = f'=SUMIFS(\'Time Entries\'!$M:$M,\'Time Entries\'!$B:$B,$B{row},\'Time Entries\'!$F:$F,">="&$D{row},\'Time Entries\'!$F:$F,"<="&$E{row},\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected")'
        ws[f"K{row}"] = f'=SUMIFS(\'Time Entries\'!$N:$N,\'Time Entries\'!$B:$B,$B{row},\'Time Entries\'!$F:$F,">="&$D{row},\'Time Entries\'!$F:$F,"<="&$E{row},\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected")'
        ws[f"L{row}"] = f'=I{row}+J{row}-K{row}'
    add_validation(ws, "Pay Period Status", "F2:F100")
    add_validation(ws, "Payment Status", "G2:G100")
    add_range_validation(ws, "Workers!$A$2:$A$200", "B2:B100")
    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].value = as_date(ws[f"D{row}"].value)
        ws[f"E{row}"].value = as_date(ws[f"E{row}"].value)
        ws[f"M{row}"].value = as_datetime(ws[f"M{row}"].value)
        ws[f"N{row}"].value = as_datetime(ws[f"N{row}"].value)
    for row in range(2, 101):
        ws[f"D{row}"].number_format = "m/d/yyyy"
        ws[f"E{row}"].number_format = "m/d/yyyy"
        ws[f"M{row}"].number_format = "m/d/yyyy h:mm"
        ws[f"N{row}"].number_format = "m/d/yyyy h:mm"
        ws[f"H{row}"].number_format = "0.00"
        for col in ["I", "J", "K", "L"]:
            ws[f"{col}{row}"].number_format = "$#,##0.00"
    ws.conditional_formatting.add("A2:O100", FormulaRule(formula=['AND($F2="Finalized",$G2="Unpaid")'], fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add("A2:O100", FormulaRule(formula=['$G2="Paid"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A2:O100", FormulaRule(formula=['$F2="Open"'], fill=PatternFill("solid", fgColor=THEME["blue"])))
    metric_card(ws, 1, 17, "Pay Period", '=IFERROR(INDEX($A$2:$A$100,MATCH("Open",$F$2:$F$100,0)),"Demo")', 2)
    metric_card(ws, 1, 19, "Approved Hours", '=SUM($H$2:$H$100)', 2, "0.00")
    metric_card(ws, 4, 17, "Worker Count", '=COUNTA(UNIQUE($B$2:$B$100))', 2)
    metric_card(ws, 4, 19, "Paid Status", '=COUNTIF($G$2:$G$100,"Paid")&" paid"', 2)
    metric_card(ws, 7, 17, "Export Batch ID", '=IFERROR(INDEX(\'Proof Exports\'!$A$2:$A$100,1),"No export")', 4)


def build_worker_proof(ws):
    ws.title = "Worker Proof"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "CrewPay Ledger Worker Proof"
    ws["A1"].font = Font(bold=True, size=20, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws.merge_cells("A1:M1")
    for col in range(1, 14):
        ws.cell(1, col).fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["A2"] = "Printable selected-worker proof packet. Use one Worker ID and one Pay Period ID only."
    ws["A2"].font = Font(italic=True, color=THEME["muted"])
    ws["A2"].fill = PatternFill("solid", fgColor=THEME["blue_2"])
    ws.merge_cells("A2:M2")
    labels = [
        ("A3", "Selected Worker ID"), ("A4", "Selected Pay Period ID"), ("A5", "Worker Name"),
        ("A6", "Worker Status"), ("A7", "Date Range"), ("A8", "Payment Status"), ("A9", "Generated Timestamp"),
        ("A10", "Proof Selector Check")
    ]
    for cell, label in labels:
        ws[cell] = label
        ws[cell].font = Font(bold=True)
    ws["B3"] = "W-1001"
    ws["B4"] = "P-4001"
    ws["B5"] = '=IFERROR(VLOOKUP($B$3,Workers!$A:$H,2,FALSE),"")'
    ws["B6"] = '=IFERROR(VLOOKUP($B$3,Workers!$A:$H,5,FALSE),"")'
    ws["B7"] = '=IFERROR(TEXT(VLOOKUP($B$4,\'Pay Periods\'!$A:$O,4,FALSE),"m/d/yyyy")&" to "&TEXT(VLOOKUP($B$4,\'Pay Periods\'!$A:$O,5,FALSE),"m/d/yyyy"),"")'
    ws["B8"] = '=IFERROR(VLOOKUP($B$4,\'Pay Periods\'!$A:$O,7,FALSE),"")'
    ws["B9"] = '=NOW()'
    ws["B10"] = '=IFERROR(IF(VLOOKUP($B$4,\'Pay Periods\'!$A:$O,2,FALSE)=$B$3,"OK - worker/pay period match","CHECK SELECTION - pay period belongs to another worker"),"CHECK SELECTION")'
    ws["B9"].number_format = "m/d/yyyy h:mm"
    ws["D3"] = "Proof Status"
    ws["E3"] = '=IF($B$10="OK - worker/pay period match","Ready for export","Check selector")'
    ws["D4"] = "Export Batch"
    ws["E4"] = '=IFERROR(INDEX(\'Proof Exports\'!$A$2:$A$100,MATCH($B$4,\'Proof Exports\'!$D$2:$D$100,0)),"Not exported")'
    ws["D5"] = "Paid Date"
    ws["E5"] = '=IFERROR(VLOOKUP($B$4,\'Pay Periods\'!$A:$O,14,FALSE),"")'
    ws["D6"] = "Archive Rule"
    ws["E6"] = "Inactive workers keep historical proof."
    for row in range(3, 7):
        ws[f"D{row}"].font = Font(bold=True, color=THEME["navy"])
        ws[f"D{row}"].fill = PatternFill("solid", fgColor=THEME["blue_2"])
        ws[f"E{row}"].fill = PatternFill("solid", fgColor=THEME["card"])
        ws[f"D{row}"].border = THIN_BORDER
        ws[f"E{row}"].border = THIN_BORDER
    ws["E5"].number_format = "m/d/yyyy h:mm"
    add_validation(ws, "Payment Status", "B8")
    dv_worker = DataValidation(type="list", formula1="Workers!$A$2:$A$200", allow_blank=False)
    dv_period = DataValidation(type="list", formula1="'Pay Periods'!$A$2:$A$100", allow_blank=False)
    ws.add_data_validation(dv_worker); dv_worker.add("B3")
    ws.add_data_validation(dv_period); dv_period.add("B4")
    headers = ["Entry ID", "Work Date", "Job ID", "Job Name", "Hours", "Rate", "Gross Pay", "Reimbursement", "Deduction", "Net Pay", "Approval Status", "Correction Note", "Notes"]
    for idx, header in enumerate(headers, 1):
        ws.cell(12, idx, header)
    header_style(ws, 12, 1, len(headers))
    ws["A13"] = '=IF($B$10<>"OK - worker/pay period match","Check selected worker/pay period before printing proof",FILTER({\'Time Entries\'!A2:A200,\'Time Entries\'!F2:F200,\'Time Entries\'!D2:D200,\'Time Entries\'!E2:E200,\'Time Entries\'!J2:J200,\'Time Entries\'!K2:K200,\'Time Entries\'!L2:L200,\'Time Entries\'!M2:M200,\'Time Entries\'!N2:N200,\'Time Entries\'!O2:O200,\'Time Entries\'!P2:P200,\'Time Entries\'!S2:S200,\'Time Entries\'!T2:T200},\'Time Entries\'!B2:B200=$B$3,\'Time Entries\'!F2:F200>=VLOOKUP($B$4,\'Pay Periods\'!$A:$O,4,FALSE),\'Time Entries\'!F2:F200<=VLOOKUP($B$4,\'Pay Periods\'!$A:$O,5,FALSE)))'
    total_row = 30
    totals = [
        ("A30", "Total Hours", "B30", '=IF($B$10<>"OK - worker/pay period match",0,SUMIFS(\'Time Entries\'!$J:$J,\'Time Entries\'!$B:$B,$B$3,\'Time Entries\'!$F:$F,">="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,4,FALSE),\'Time Entries\'!$F:$F,"<="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,5,FALSE),\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected"))'),
        ("A31", "Gross Pay", "B31", '=IF($B$10<>"OK - worker/pay period match",0,SUMIFS(\'Time Entries\'!$L:$L,\'Time Entries\'!$B:$B,$B$3,\'Time Entries\'!$F:$F,">="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,4,FALSE),\'Time Entries\'!$F:$F,"<="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,5,FALSE),\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected"))'),
        ("A32", "Reimbursements", "B32", '=IF($B$10<>"OK - worker/pay period match",0,SUMIFS(\'Time Entries\'!$M:$M,\'Time Entries\'!$B:$B,$B$3,\'Time Entries\'!$F:$F,">="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,4,FALSE),\'Time Entries\'!$F:$F,"<="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,5,FALSE),\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected"))'),
        ("A33", "Deductions", "B33", '=IF($B$10<>"OK - worker/pay period match",0,SUMIFS(\'Time Entries\'!$N:$N,\'Time Entries\'!$B:$B,$B$3,\'Time Entries\'!$F:$F,">="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,4,FALSE),\'Time Entries\'!$F:$F,"<="&VLOOKUP($B$4,\'Pay Periods\'!$A:$O,5,FALSE),\'Time Entries\'!$P:$P,"<>Draft",\'Time Entries\'!$P:$P,"<>Rejected"))'),
        ("A34", "Net Pay", "B34", '=B31+B32-B33'),
    ]
    for label_cell, label, value_cell, formula in totals:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = formula
    for row in range(31, 35):
        ws[f"B{row}"].number_format = "$#,##0.00"
    ws["B30"].number_format = "0.00"
    card(ws, 8, 4, 3, 4, "Correction / Dispute Notes", '=IFERROR(TEXTJOIN(CHAR(10),TRUE,FILTER(\'Correction Log\'!$H$2:$H$100,\'Correction Log\'!$C$2:$C$100=$B$3,\'Correction Log\'!$E$2:$E$100=$B$4)),"No correction notes for selected proof.")', THEME["card"])
    card(ws, 8, 9, 3, 5, "Read-only Archive Note", "Former or inactive workers can keep proof access for historical records, but should not receive new time entries.", THEME["blue_2"])
    table_style(ws, 3, 10, 1, 2)
    ws.conditional_formatting.add("A10:B10", FormulaRule(formula=['$B$10<>"OK - worker/pay period match"'], fill=PatternFill("solid", fgColor="F8D7DA")))
    ws.conditional_formatting.add("D3:E6", FormulaRule(formula=['$E$3="Ready for export"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A13:M29", FormulaRule(formula=['$K13="Paid"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A13:M29", FormulaRule(formula=['$K13="Approved"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A13:M29", FormulaRule(formula=['$K13="Submitted"'], fill=PatternFill("solid", fgColor=THEME["yellow"])))
    set_widths(ws, {"A": 18, "B": 22, "C": 14, "D": 24, "E": 12, "F": 12, "G": 14, "H": 16, "I": 12, "J": 14, "K": 18, "L": 28, "M": 34})
    ws.freeze_panes = "A12"
    ws.print_title_rows = "1:12"
    ws.print_area = "A1:M34"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_dashboard(ws):
    ws.title = "Dashboard"
    title_bar(ws, "CrewPay Ledger Dashboard", "Workbook-first operating snapshot for worker proof, approvals, pay periods, and audit readiness.", 10)
    metrics = [
        (4, 1, "Active Workers", '=COUNTIF(Workers!$E$2:$E$200,"Active")', None),
        (4, 3, "Submitted Entries", '=COUNTIF(\'Time Entries\'!$P$2:$P$200,"Submitted")', None),
        (4, 5, "Approved Entries", '=COUNTIF(\'Time Entries\'!$P$2:$P$200,"Approved")+COUNTIF(\'Time Entries\'!$P$2:$P$200,"Finalized")+COUNTIF(\'Time Entries\'!$P$2:$P$200,"Paid")', None),
        (4, 7, "Rejected Entries", '=COUNTIF(\'Time Entries\'!$P$2:$P$200,"Rejected")', None),
        (8, 1, "Current Pay Period", '=IFERROR(INDEX(\'Pay Periods\'!$A$2:$A$100,MATCH("Open",\'Pay Periods\'!$F$2:$F$100,0)),"Demo sample")', None),
        (8, 3, "Total Approved Hours", '=SUMIFS(\'Time Entries\'!$J$2:$J$200,\'Time Entries\'!$P$2:$P$200,"Approved")+SUMIFS(\'Time Entries\'!$J$2:$J$200,\'Time Entries\'!$P$2:$P$200,"Finalized")+SUMIFS(\'Time Entries\'!$J$2:$J$200,\'Time Entries\'!$P$2:$P$200,"Paid")', "0.00"),
        (8, 5, "Paid Proof Records", '=COUNTIFS(\'Proof Exports\'!$E$2:$E$100,"<>",\'Proof Exports\'!$D$2:$D$100,"<>")', None),
        (8, 7, "Inactive Workers", '=COUNTIF(Workers!$E$2:$E$200,"Inactive")', None),
        (12, 1, "Corrections Pending", '=COUNTIF(\'Time Entries\'!$S$2:$S$200,"<>")', None),
    ]
    for row, col, label, formula, number_format in metrics:
        metric_card(ws, row, col, label, formula, 2, number_format)
    card(ws, 12, 3, 2, 4, "Proof Privacy Rule", "Worker Proof and exports are selected-worker / selected-pay-period only.", THEME["blue_2"])
    card(ws, 12, 7, 2, 3, "Schedule Boundary", "Calendar is a planning mirror. It never becomes proof.", THEME["blue_2"])
    ws["A16"] = "Operational Status"
    ws["A16"].font = Font(bold=True, size=13, color=THEME["navy"])
    status_rows = [
        ("Worker access", '=IF(COUNTIF(Workers!$E$2:$E$200,"Inactive")>0,"Inactive archive present","All active")'),
        ("Approval queue", '=IF(COUNTIF(\'Time Entries\'!$P$2:$P$200,"Submitted")>0,"Review submitted time","No submitted entries")'),
        ("Proof exports", '=IF(COUNTA(\'Proof Exports\'!$A$2:$A$100)>0,"Export log active","No exports logged")'),
        ("Corrections", '=IF(COUNTA(\'Correction Log\'!$A$2:$A$100)>0,"Correction log active","No corrections logged")'),
    ]
    for idx, (label, formula) in enumerate(status_rows, 17):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, formula)
    header_style(ws, 16, 1, 2)
    table_style(ws, 16, 20, 1, 2)
    card(ws, 16, 5, 5, 5, "Demo Boundaries", "No worker accounts, backend database, payroll tax logic, HR compliance workflow, chat, or worker-to-worker messaging are part of this workbook demo.", THEME["card"])
    set_widths(ws, {"A": 20, "B": 14, "C": 20, "D": 14, "E": 20, "F": 14, "G": 20, "H": 14, "I": 14, "J": 14})
    ws.freeze_panes = "A3"


def build_access_status_demo(ws):
    ws.title = "Access Status Demo"
    title_bar(ws, "Access Status Demo", "Visual guide to future-use access versus historical proof access.", 8)
    scenarios = [
        (4, 1, "Active worker", "Can submit new time\nCan appear on Schedule\nCan receive admin notices", THEME["green"]),
        (4, 4, "Inactive worker", "Cannot receive new time entries\nHistorical rows stay visible\nProof can still be regenerated", THEME["gray"]),
        (4, 7, "Former / read-only", "No future-use workflow\nProof archive remains available\nAdmin should not delete records", THEME["blue_2"]),
    ]
    for row, col, heading, body, fill in scenarios:
        card(ws, row, col, 4, 2, heading, body, fill)
    matrix_headers = ["Rule", "Active", "Inactive", "Former / Read-only", "Admin Note"]
    for col, header in enumerate(matrix_headers, 1):
        ws.cell(10, col, header)
    matrix = [
        ["Create new time entry", "Yes", "No", "No", "Inactive means future-use access is stopped."],
        ["View historical proof", "Yes", "Yes", "Yes", "Proof protection survives status changes."],
        ["Appear in current schedule", "Yes", "No", "No", "Schedule is planning, not proof."],
        ["Keep correction history", "Yes", "Yes", "Yes", "Corrections stay visible, not silent."],
    ]
    for row in matrix:
        ws.append(row)
    header_style(ws, 10, 1, len(matrix_headers))
    table_style(ws, 10, 14, 1, len(matrix_headers))
    ws.conditional_formatting.add("A11:E14", FormulaRule(formula=['B11="Yes"'], fill=PatternFill("solid", fgColor=THEME["green"])))
    ws.conditional_formatting.add("A11:E14", FormulaRule(formula=['C11="No"'], fill=PatternFill("solid", fgColor=THEME["gray"])))
    card(ws, 17, 1, 3, 7, "Demo takeaway", "Admin can stop future access, but should not delete historical rows. Worker proof remains tied to workbook records.", THEME["card"])
    set_widths(ws, {"A": 24, "B": 14, "C": 16, "D": 20, "E": 48, "F": 14, "G": 18, "H": 18})
    ws.freeze_panes = "A10"


def build_workflow_demo(ws):
    ws.title = "Workflow Demo"
    title_bar(ws, "Workflow Demo", "One-screen view of the CrewPay Ledger operating flow.", 12)
    steps = [
        (4, 1, "1", "Worker enters time", "Draft row in Time Entries"),
        (4, 4, "2", "Worker submits", "Status becomes Submitted"),
        (4, 7, "3", "Admin reviews", "Approve, reject, or request correction"),
        (4, 10, "4", "Approve / Reject", "Decision remains visible"),
        (10, 1, "5", "Pay period created", "Totals summarize approved time"),
        (10, 4, "6", "Worker Proof available", "Selected worker/pay period only"),
        (10, 7, "7", "Proof exported", "PDF/CSV export logged"),
        (10, 10, "8", "Read-only archive", "Inactive workers keep proof history"),
    ]
    for row, col, num, heading, detail in steps:
        card(ws, row, col, 4, 2, f"{num}. {heading}", detail, THEME["card"])
    arrows = [(4, 3), (4, 6), (4, 9), (8, 10), (10, 3), (10, 6), (10, 9)]
    for row, col in arrows:
        ws.cell(row + 1, col, "->")
        ws.cell(row + 1, col).font = Font(bold=True, size=18, color=THEME["navy"])
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
    card(ws, 16, 1, 3, 11, "Source-of-truth rule", "Schedule, Admin Notices, Calendar sync, and export files are helper/reference layers. Time Entries, Pay Periods, Correction Log, and Worker Proof remain the proof trail.", THEME["blue_2"])
    set_widths(ws, {get_column_letter(i): 16 for i in range(1, 13)})
    for row in range(1, 20):
        ws.row_dimensions[row].height = 26


def build_simple_logs(wb):
    append_table(wb["Proof Exports"], ["Export ID", "Worker ID", "Worker Name", "Pay Period ID", "Export Type", "Generated At", "Generated By", "Export Reference", "Notes"], PROOF_EXPORTS, {"A": 12, "B": 12, "C": 20, "D": 15, "E": 12, "F": 18, "G": 16, "H": 24, "I": 44})
    add_range_validation(wb["Proof Exports"], "Workers!$A$2:$A$200", "B2:B100")
    add_range_validation(wb["Proof Exports"], "'Pay Periods'!$A$2:$A$100", "D2:D100")
    add_validation(wb["Proof Exports"], "Export Type", "E2:E100")
    add_status_formatting(wb["Proof Exports"], "A2:I100", "E")
    append_table(wb["Access Log"], ["Log ID", "Worker ID", "Worker Name", "Previous Status", "New Status", "Changed At", "Changed By", "Reason"], ACCESS_LOG, {"A": 12, "B": 12, "C": 20, "D": 16, "E": 16, "F": 18, "G": 16, "H": 46})
    add_range_validation(wb["Access Log"], "Workers!$A$2:$A$200", "B2:B100")
    add_validation(wb["Access Log"], "Access Status", "D2:E100")
    add_status_formatting(wb["Access Log"], "A2:H100", "E")
    append_table(wb["Correction Log"], ["Correction ID", "Entry ID", "Worker ID", "Worker Name", "Pay Period ID", "Correction Date", "Corrected By", "Correction Reason", "Original Value Summary", "New Value Summary", "Notes"], CORRECTION_LOG, {"A": 14, "B": 12, "C": 12, "D": 20, "E": 15, "F": 15, "G": 16, "H": 24, "I": 24, "J": 24, "K": 38})
    add_range_validation(wb["Correction Log"], "'Time Entries'!$A$2:$A$200", "B2:B100")
    add_range_validation(wb["Correction Log"], "Workers!$A$2:$A$200", "C2:C100")
    add_range_validation(wb["Correction Log"], "'Pay Periods'!$A$2:$A$100", "E2:E100")
    card(wb["Correction Log"], 1, 13, 3, 3, "Correction rule", "Corrections are visible records. Do not silently overwrite proof.", THEME["blue_2"])
    append_table(wb["Schedule"], ["Schedule ID", "Job ID", "Job Name", "Worker ID", "Worker Name", "Scheduled Date", "Start Time", "End Time", "Schedule Status", "Calendar Event ID", "Notes"], SCHEDULE, {"A": 13, "B": 12, "C": 24, "D": 12, "E": 20, "F": 15, "G": 12, "H": 12, "I": 16, "J": 24, "K": 42})
    add_range_validation(wb["Schedule"], "Jobs!$A$2:$A$200", "B2:B100")
    add_range_validation(wb["Schedule"], "Workers!$A$2:$A$200", "D2:D100")
    add_validation(wb["Schedule"], "Schedule Status", "I2:I100")
    add_status_formatting(wb["Schedule"], "A2:K100", "I")
    card(wb["Schedule"], 1, 13, 3, 3, "Schedule boundary", "Planning and Calendar reference only. Not proof.", THEME["blue_2"])
    append_table(wb["Admin Notices"], ["Notice ID", "Created At", "Created By", "Recipient Type", "Worker ID", "Worker Name", "Subject", "Message", "Related Pay Period ID", "Delivery Method", "Notice Status", "Sent At", "Notes"], ADMIN_NOTICES, {"A": 12, "B": 18, "C": 16, "D": 20, "E": 12, "F": 20, "G": 24, "H": 44, "I": 20, "J": 16, "K": 15, "L": 18, "M": 38})
    add_validation(wb["Admin Notices"], "Recipient Type", "D2:D100")
    add_range_validation(wb["Admin Notices"], "Workers!$A$2:$A$200", "E2:E100")
    add_range_validation(wb["Admin Notices"], "'Pay Periods'!$A$2:$A$100", "I2:I100")
    add_validation(wb["Admin Notices"], "Delivery Method", "J2:J100")
    add_validation(wb["Admin Notices"], "Notice Status", "K2:K100")
    add_status_formatting(wb["Admin Notices"], "A2:M100", "K")
    card(wb["Admin Notices"], 1, 15, 3, 3, "Notice boundary", "One-way admin notices only. Not chat and not proof.", THEME["blue_2"])
    append_table(wb["Calendar Sync Log"], ["Calendar Log ID", "Job ID", "Job Name", "Worker ID", "Worker Name", "Calendar Event ID", "Event Date", "Sync Status", "Last Synced At", "Notes"], CALENDAR_SYNC, {"A": 16, "B": 12, "C": 24, "D": 12, "E": 20, "F": 24, "G": 14, "H": 16, "I": 18, "J": 44})
    add_range_validation(wb["Calendar Sync Log"], "Jobs!$A$2:$A$200", "B2:B100")
    add_range_validation(wb["Calendar Sync Log"], "Workers!$A$2:$A$200", "D2:D100")
    add_validation(wb["Calendar Sync Log"], "Sync Status", "H2:H100")
    add_status_formatting(wb["Calendar Sync Log"], "A2:J100", "H")
    card(wb["Calendar Sync Log"], 1, 12, 3, 3, "Calendar boundary", "Calendar mirrors schedule references. It never creates proof.", THEME["blue_2"])

    for row in range(2, wb["Proof Exports"].max_row + 1):
        wb["Proof Exports"][f"F{row}"].value = as_datetime(wb["Proof Exports"][f"F{row}"].value)
        wb["Proof Exports"][f"F{row}"].number_format = "m/d/yyyy h:mm"
    for row in range(2, wb["Access Log"].max_row + 1):
        wb["Access Log"][f"F{row}"].value = as_datetime(wb["Access Log"][f"F{row}"].value)
        wb["Access Log"][f"F{row}"].number_format = "m/d/yyyy h:mm"
    for row in range(2, wb["Correction Log"].max_row + 1):
        wb["Correction Log"][f"F{row}"].value = as_date(wb["Correction Log"][f"F{row}"].value)
        wb["Correction Log"][f"F{row}"].number_format = "m/d/yyyy"
    for row in range(2, wb["Schedule"].max_row + 1):
        wb["Schedule"][f"F{row}"].value = as_date(wb["Schedule"][f"F{row}"].value)
        wb["Schedule"][f"G{row}"].value = as_time(wb["Schedule"][f"G{row}"].value)
        wb["Schedule"][f"H{row}"].value = as_time(wb["Schedule"][f"H{row}"].value)
        wb["Schedule"][f"F{row}"].number_format = "m/d/yyyy"
        wb["Schedule"][f"G{row}"].number_format = "h:mm"
        wb["Schedule"][f"H{row}"].number_format = "h:mm"
    for row in range(2, wb["Admin Notices"].max_row + 1):
        wb["Admin Notices"][f"B{row}"].value = as_datetime(wb["Admin Notices"][f"B{row}"].value)
        wb["Admin Notices"][f"L{row}"].value = as_datetime(wb["Admin Notices"][f"L{row}"].value)
        wb["Admin Notices"][f"B{row}"].number_format = "m/d/yyyy h:mm"
        wb["Admin Notices"][f"L{row}"].number_format = "m/d/yyyy h:mm"
    for row in range(2, wb["Calendar Sync Log"].max_row + 1):
        wb["Calendar Sync Log"][f"G{row}"].value = as_date(wb["Calendar Sync Log"][f"G{row}"].value)
        wb["Calendar Sync Log"][f"I{row}"].value = as_datetime(wb["Calendar Sync Log"][f"I{row}"].value)
        wb["Calendar Sync Log"][f"G{row}"].number_format = "m/d/yyyy"
        wb["Calendar Sync Log"][f"I{row}"].number_format = "m/d/yyyy h:mm"


def build_bridge_tabs(wb):
    append_table(
        wb["App Config"],
        ["Config Key", "Config Value", "Notes"],
        APP_CONFIG,
        {"A": 28, "B": 44, "C": 58},
    )
    add_validation(wb["App Config"], "Boolean Flag", "B2")
    add_validation(wb["App Config"], "Boolean Flag", "B6:B8")
    card(
        wb["App Config"],
        1,
        5,
        4,
        4,
        "Bridge boundary",
        "The workbook remains source of truth. The app may submit controlled intake records through Apps Script after deployment.",
        THEME["blue_2"],
    )

    append_table(
        wb["Pending Worker Intake"],
        ["Intake ID", "Submitted At", "Submission Source", "Submission Status", "Worker ID", "Worker Name", "Access Status", "Role / Trade", "Contact", "Notes", "Reviewed At", "Reviewed By", "Review Notes"],
        PENDING_WORKER_INTAKE,
        {"A": 13, "B": 18, "C": 18, "D": 18, "E": 12, "F": 22, "G": 15, "H": 18, "I": 24, "J": 36, "K": 18, "L": 18, "M": 36},
    )
    add_validation(wb["Pending Worker Intake"], "Submission Status", "D2:D200")
    add_validation(wb["Pending Worker Intake"], "Access Status", "G2:G200")
    add_status_formatting(wb["Pending Worker Intake"], "A2:M200", "D")

    append_table(
        wb["Pending Pay Period Intake"],
        ["Intake ID", "Submitted At", "Submission Source", "Submission Status", "Pay Period ID", "Worker ID", "Worker Name", "Period Start", "Period End", "Pay Date", "Notes", "Reviewed At", "Reviewed By", "Review Notes"],
        PENDING_PAY_PERIOD_INTAKE,
        {"A": 13, "B": 18, "C": 18, "D": 18, "E": 16, "F": 12, "G": 22, "H": 14, "I": 14, "J": 14, "K": 36, "L": 18, "M": 18, "N": 36},
    )
    add_validation(wb["Pending Pay Period Intake"], "Submission Status", "D2:D200")
    add_status_formatting(wb["Pending Pay Period Intake"], "A2:N200", "D")

    append_table(
        wb["Pending Time Entries"],
        ["Intake ID", "Submitted At", "Submission Source", "Submission Status", "Entry ID", "Worker ID", "Worker Name", "Pay Period ID", "Work Date", "Job / Work Type", "Hours", "Rate", "Amount", "Notes", "Reviewed At", "Reviewed By", "Review Notes"],
        PENDING_TIME_ENTRIES,
        {"A": 13, "B": 18, "C": 18, "D": 18, "E": 12, "F": 12, "G": 22, "H": 15, "I": 14, "J": 24, "K": 10, "L": 12, "M": 14, "N": 36, "O": 18, "P": 18, "Q": 36},
    )
    add_validation(wb["Pending Time Entries"], "Submission Status", "D2:D200")
    add_status_formatting(wb["Pending Time Entries"], "A2:Q200", "D")

    append_table(
        wb["App Submission Log"],
        ["Log ID", "Submitted At", "Action", "Submission Source", "Status", "Related Intake ID", "Related Worker ID", "Related Pay Period ID", "Message", "Raw Payload Summary", "Handled By Script Version"],
        APP_SUBMISSION_LOG,
        {"A": 12, "B": 18, "C": 24, "D": 18, "E": 14, "F": 18, "G": 18, "H": 20, "I": 34, "J": 34, "K": 24},
    )

    append_table(
        wb["Bridge Schema"],
        ["Action", "Target Tab", "Required Fields", "Optional Fields", "Success Response", "Error Conditions", "Notes"],
        BRIDGE_SCHEMA,
        {"A": 24, "B": 26, "C": 48, "D": 38, "E": 38, "F": 42, "G": 48},
    )

    for ws_name in ["Pending Worker Intake", "Pending Pay Period Intake", "Pending Time Entries", "App Submission Log"]:
        ws = wb[ws_name]
        for row in range(2, ws.max_row + 1):
            if ws.max_column >= 2 and ws[f"B{row}"].value:
                ws[f"B{row}"].value = as_datetime(ws[f"B{row}"].value)
                ws[f"B{row}"].number_format = "m/d/yyyy h:mm"
        if ws_name == "Pending Pay Period Intake":
            for row in range(2, ws.max_row + 1):
                for col in ["H", "I", "J"]:
                    ws[f"{col}{row}"].value = as_date(ws[f"{col}{row}"].value)
                    ws[f"{col}{row}"].number_format = "m/d/yyyy"
        if ws_name == "Pending Time Entries":
            for row in range(2, ws.max_row + 1):
                ws[f"I{row}"].value = as_date(ws[f"I{row}"].value)
                ws[f"I{row}"].number_format = "m/d/yyyy"
                ws[f"K{row}"].number_format = "0.00"
                ws[f"L{row}"].number_format = "$#,##0.00"
                ws[f"M{row}"].number_format = "$#,##0.00"


def build_workbook():
    wb = Workbook()
    for sheet in SHEET_ORDER[1:]:
        wb.create_sheet(sheet)
    wb._sheets = [wb[name] for name in wb.sheetnames if name in SHEET_ORDER]
    wb._sheets = [wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet) for sheet in SHEET_ORDER]
    build_instructions(wb["Instructions"])
    build_dashboard(wb["Dashboard"])
    build_workers(wb["Workers"])
    build_jobs(wb["Jobs"])
    build_time_entries(wb["Time Entries"])
    build_pay_periods(wb["Pay Periods"])
    build_worker_proof(wb["Worker Proof"])
    build_access_status_demo(wb["Access Status Demo"])
    build_workflow_demo(wb["Workflow Demo"])
    build_simple_logs(wb)
    build_bridge_tabs(wb)
    build_dropdowns(wb["Dropdown Lists"])

    # Format date/time columns in simple logs.
    for ws_name in ["Proof Exports", "Access Log", "Correction Log", "Schedule", "Admin Notices", "Calendar Sync Log"]:
        table_style(wb[ws_name], 1, wb[ws_name].max_row, 1, wb[ws_name].max_column)
    wb.save(Path(__file__).resolve().parent / OUTPUT_FILE)
    return Path(__file__).resolve().parent / OUTPUT_FILE


def has_formula(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                return True
    return False


def audit(path):
    wb = load_workbook(path, data_only=False)
    required = {
        "Workers": ["Worker ID", "Worker Name", "Worker Email", "Role", "Access Status", "Created At", "Inactive At", "Notes"],
        "Jobs": ["Job ID", "Job Name", "Client or Site", "Status", "Default Rate", "Calendar Event ID", "Notes"],
        "Time Entries": ["Entry ID", "Worker ID", "Worker Name", "Job ID", "Job Name", "Work Date", "Start Time", "End Time", "Break Minutes", "Hours", "Rate", "Gross Pay", "Reimbursement", "Deduction", "Net Pay", "Approval Status", "Submitted At", "Approved At", "Correction Note", "Notes"],
        "Pay Periods": ["Pay Period ID", "Worker ID", "Worker Name", "Period Start", "Period End", "Status", "Payment Status", "Total Hours", "Gross Pay", "Reimbursement Total", "Deduction Total", "Net Pay", "Finalized At", "Paid At", "Notes"],
        "Proof Exports": ["Export ID", "Worker ID", "Worker Name", "Pay Period ID", "Export Type", "Generated At", "Generated By", "Export Reference", "Notes"],
        "Access Log": ["Log ID", "Worker ID", "Worker Name", "Previous Status", "New Status", "Changed At", "Changed By", "Reason"],
        "Correction Log": ["Correction ID", "Entry ID", "Worker ID", "Worker Name", "Pay Period ID", "Correction Date", "Corrected By", "Correction Reason", "Original Value Summary", "New Value Summary", "Notes"],
        "Schedule": ["Schedule ID", "Job ID", "Job Name", "Worker ID", "Worker Name", "Scheduled Date", "Start Time", "End Time", "Schedule Status", "Calendar Event ID", "Notes"],
        "Admin Notices": ["Notice ID", "Created At", "Created By", "Recipient Type", "Worker ID", "Worker Name", "Subject", "Message", "Related Pay Period ID", "Delivery Method", "Notice Status", "Sent At", "Notes"],
        "Calendar Sync Log": ["Calendar Log ID", "Job ID", "Job Name", "Worker ID", "Worker Name", "Calendar Event ID", "Event Date", "Sync Status", "Last Synced At", "Notes"],
        "App Config": ["Config Key", "Config Value", "Notes"],
        "Pending Worker Intake": ["Intake ID", "Submitted At", "Submission Source", "Submission Status", "Worker ID", "Worker Name", "Access Status", "Role / Trade", "Contact", "Notes", "Reviewed At", "Reviewed By", "Review Notes"],
        "Pending Pay Period Intake": ["Intake ID", "Submitted At", "Submission Source", "Submission Status", "Pay Period ID", "Worker ID", "Worker Name", "Period Start", "Period End", "Pay Date", "Notes", "Reviewed At", "Reviewed By", "Review Notes"],
        "Pending Time Entries": ["Intake ID", "Submitted At", "Submission Source", "Submission Status", "Entry ID", "Worker ID", "Worker Name", "Pay Period ID", "Work Date", "Job / Work Type", "Hours", "Rate", "Amount", "Notes", "Reviewed At", "Reviewed By", "Review Notes"],
        "App Submission Log": ["Log ID", "Submitted At", "Action", "Submission Source", "Status", "Related Intake ID", "Related Worker ID", "Related Pay Period ID", "Message", "Raw Payload Summary", "Handled By Script Version"],
        "Bridge Schema": ["Action", "Target Tab", "Required Fields", "Optional Fields", "Success Response", "Error Conditions", "Notes"],
        "Dropdown Lists": list(DROPDOWNS.keys()),
    }
    checks = [
        ("workbook file exists", path.exists()),
        ("exactly 22 tabs", len(wb.sheetnames) == 22),
        ("tab order matches prompt", wb.sheetnames == SHEET_ORDER),
        ("Dashboard formulas exist", has_formula(wb["Dashboard"], "A5:H8")),
        ("Time Entries formulas exist", has_formula(wb["Time Entries"], "C2:O200")),
        ("Pay Period formulas exist", has_formula(wb["Pay Periods"], "C2:L100")),
        ("Worker Proof formulas exist", has_formula(wb["Worker Proof"], "A13:M34")),
        ("Worker Proof selected worker selector exists", wb["Worker Proof"]["B3"].value == "W-1001"),
        ("Worker Proof selector check exists", "worker/pay period match" in str(wb["Worker Proof"]["B10"].value)),
        ("Worker Proof print area exists", bool(wb["Worker Proof"].print_area)),
        ("Schedule tab states not proof", "not proof" in str(wb["Schedule"]["K2"].value).lower()),
        ("Admin Notices tab states not chat/proof", "not chat" in str(wb["Admin Notices"]["M2"].value).lower()),
        ("active worker sample exists", any(row[4].value == "Active" for row in wb["Workers"].iter_rows(min_row=2, max_col=5))),
        ("inactive worker sample exists", any(row[4].value == "Inactive" for row in wb["Workers"].iter_rows(min_row=2, max_col=5))),
    ]
    for sheet_name, headers in required.items():
        actual = [wb[sheet_name].cell(1, col).value for col in range(1, len(headers) + 1)]
        checks.append((f"{sheet_name} required headers exist", actual == headers))
    validation_count = sum(len(ws.data_validations.dataValidation) for ws in wb.worksheets)
    checks.append(("data validation applied", validation_count >= 24))
    text = " ".join(str(cell.value).lower() for ws in wb.worksheets for row in ws.iter_rows() for cell in row if cell.value)
    forbidden = ["gmail.com", "outlook.com", "yahoo.com", "api key", "password", "secret=", "bearer "]
    checks.append(("no real/private data strings", not any(term in text for term in forbidden)))
    checks.append(("App Config bridge disabled by default", wb["App Config"]["B6"].value == "FALSE"))
    checks.append(("Pending Time Entries amount formula exists", str(wb["Pending Time Entries"]["M2"].value).startswith("=")))
    passed = all(result for _, result in checks)
    print("CrewPay Ledger workbook audit")
    print(f"Output file path: {path}")
    print(f"Sheet names: {wb.sheetnames}")
    for label, result in checks:
        print(f"{'PASS' if result else 'FAIL'} - {label}")
    print(f"Overall audit: {'PASS' if passed else 'FAIL'}")
    return passed


if __name__ == "__main__":
    output = build_workbook()
    if not audit(output):
        raise SystemExit(1)
