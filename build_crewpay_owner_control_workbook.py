from __future__ import annotations

from datetime import date
from pathlib import Path
import sys


LOCAL_DEPS = Path(__file__).resolve().parent / ".deps" / "usr" / "lib" / "python3" / "dist-packages"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "CrewPay_Owner_Control_Workbook.xlsx"

SHEET_ORDER = [
    "Instructions",
    "Owner Dashboard",
    "Client Registry",
    "Client Access Control",
    "License Billing",
    "Bridge Registry",
    "Feature Flags",
    "System Health",
    "Calendar Visibility",
    "Support Notes",
    "Owner Audit Log",
    "Dropdown Lists",
    "Data Dictionary",
    "Apps Script Setup",
]

TAB_HEADERS = {
    "Instructions": ["Section", "Details", "Privacy Boundary"],
    "Owner Dashboard": ["Operating Note", "Details"],
    "Client Registry": [
        "Client ID",
        "Client Display Name",
        "Client Legal Name",
        "Client Status",
        "Primary Contact Alias",
        "Contact Method",
        "Region",
        "Industry",
        "Onboarded Date",
        "Offboarded Date",
        "Client Ledger Location Alias",
        "Worker Data Owner",
        "Allowed Worker Count",
        "Current Worker Count Reported",
        "Data Privacy Level",
        "Notes",
    ],
    "Client Access Control": [
        "Access Record ID",
        "Client ID",
        "Client Display Name",
        "Access Status",
        "Access Start Date",
        "Access End Date",
        "Access Reason",
        "Disabled New Submissions",
        "Disabled Bridge",
        "Disabled Calendar Sync",
        "Last Access Review",
        "Reviewed By Alias",
        "Notes",
    ],
    "License Billing": [
        "License Record ID",
        "Client ID",
        "Plan Tier",
        "Billing Status",
        "Billing Period",
        "Renewal Date",
        "Allowed Worker Count",
        "Billing Worker Count",
        "Last Invoice Alias",
        "Payment Method Alias",
        "Grace Period Ends",
        "Billing Notes",
    ],
    "Bridge Registry": [
        "Bridge Record ID",
        "Client ID",
        "Bridge Status",
        "Bridge Endpoint Alias",
        "Token Status",
        "Token Last Rotated",
        "Token Rotation Due",
        "Last Health Check",
        "Last Successful Submit",
        "Last Failed Submit",
        "Pending Intake Count",
        "Last Error Summary",
        "Notes",
    ],
    "Feature Flags": [
        "Flag Record ID",
        "Client ID",
        "Feature Name",
        "Feature Status",
        "Effective Date",
        "Expiration Date",
        "Requires Billing Good Standing",
        "Requires Bridge Healthy",
        "Notes",
    ],
    "System Health": [
        "Health Record ID",
        "Client ID",
        "Check Date",
        "Ledger Status",
        "Bridge Status",
        "Worker App Status",
        "Calendar Status",
        "Backup Status",
        "Last Backup Alias",
        "Issue Severity",
        "Issue Summary",
        "Owner Action Needed",
        "Resolved Date",
        "Notes",
    ],
    "Calendar Visibility": [
        "Calendar Rule ID",
        "Client ID",
        "Calendar Visibility Status",
        "Allowed Calendar Use",
        "Worker Detail Exposure",
        "Schedule Detail Exposure",
        "Proof Source",
        "Sync Direction",
        "Last Calendar Sync",
        "Notes",
    ],
    "Support Notes": [
        "Support Note ID",
        "Client ID",
        "Note Date",
        "Note Type",
        "Priority",
        "Status",
        "Summary",
        "Owner Next Action",
        "Follow-up Date",
        "Resolved Date",
        "Notes",
    ],
    "Owner Audit Log": [
        "Audit ID",
        "Timestamp",
        "Actor Alias",
        "Area",
        "Client ID",
        "Action",
        "Previous Value",
        "New Value",
        "Reason",
        "Notes",
    ],
    "Dropdown Lists": [
        "Client Status",
        "Access Status",
        "Billing Status",
        "Plan Tier",
        "Bridge Status",
        "Token Status",
        "Feature Status",
        "Calendar Visibility Status",
        "Calendar Status",
        "Worker Detail Exposure",
        "Schedule Detail Exposure",
        "Proof Source",
        "Sync Direction",
        "Priority",
        "Support Status",
        "Issue Severity",
        "Boolean",
        "Worker Data Owner",
        "Data Privacy Level",
        "Ledger Status",
        "Worker App Status",
        "Backup Status",
    ],
    "Data Dictionary": ["Tab", "Major Fields", "Privacy Boundary", "Notes"],
    "Apps Script Setup": ["Step", "Instructions", "Notes"],
}

WORKBOOK_TITLE = "CrewPay Owner Control Workbook"
WORKBOOK_SUBTITLE = "Tim-only private control plane for client status, licensing, bridge health, feature flags, and support."

THEME = {
    "navy": "17324D",
    "blue": "DCEBF7",
    "blue_2": "EEF6FC",
    "teal": "DCEFE9",
    "green": "D9EAD3",
    "green_text": "166534",
    "yellow": "FFF2CC",
    "orange": "FCE4D6",
    "red": "F8D7DA",
    "gray": "E5E7EB",
    "gray_2": "F8FAFC",
    "muted": "6B7280",
    "text": "1F2937",
    "white": "FFFFFF",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=THEME["gray"]),
    right=Side(style="thin", color=THEME["gray"]),
    top=Side(style="thin", color=THEME["gray"]),
    bottom=Side(style="thin", color=THEME["gray"]),
)

SAMPLE_CLIENTS = [
    {
        "client_id": "CL-001",
        "display_name": "Demo Field Services",
        "legal_name": "Demo Field Services LLC",
        "status": "Active",
        "primary_contact_alias": "demo.owner.alias",
        "contact_method": "Support Alias",
        "region": "Central",
        "industry": "Field Services",
        "onboarded": "2026-05-01",
        "offboarded": "",
        "ledger_location_alias": "LEDGER-ALIAS-DEMO-001",
        "worker_data_owner": "Client Owned",
        "allowed_worker_count": 14,
        "reported_worker_count": 12,
        "privacy": "Client Controlled",
        "notes": "Primary demo client for Tim-only control examples.",
    },
    {
        "client_id": "CL-002",
        "display_name": "Sample Industrial Client",
        "legal_name": "Sample Industrial Client LLC",
        "status": "Paused",
        "primary_contact_alias": "sample.billing.alias",
        "contact_method": "Billing Alias",
        "region": "West",
        "industry": "Industrial",
        "onboarded": "2026-05-08",
        "offboarded": "",
        "ledger_location_alias": "LEDGER-ALIAS-DEMO-002",
        "worker_data_owner": "Client Owned",
        "allowed_worker_count": 8,
        "reported_worker_count": 8,
        "privacy": "Count Only",
        "notes": "Paused during plan review.",
    },
    {
        "client_id": "CL-003",
        "display_name": "Training Utility Client",
        "legal_name": "Training Utility Client Co.",
        "status": "Suspended",
        "primary_contact_alias": "training.support.alias",
        "contact_method": "Support Alias",
        "region": "South",
        "industry": "Utilities",
        "onboarded": "2026-05-15",
        "offboarded": "",
        "ledger_location_alias": "LEDGER-ALIAS-DEMO-003",
        "worker_data_owner": "Client Owned",
        "allowed_worker_count": 20,
        "reported_worker_count": 19,
        "privacy": "Alias Only",
        "notes": "Suspended pending owner review.",
    },
    {
        "client_id": "CL-004",
        "display_name": "Pilot Construction Client",
        "legal_name": "Pilot Construction Client LLC",
        "status": "Offboarded",
        "primary_contact_alias": "pilot.archive.alias",
        "contact_method": "Registry Alias",
        "region": "North",
        "industry": "Construction",
        "onboarded": "2026-04-10",
        "offboarded": "2026-06-02",
        "ledger_location_alias": "LEDGER-ALIAS-DEMO-004",
        "worker_data_owner": "Client Exported Summary",
        "allowed_worker_count": 6,
        "reported_worker_count": 0,
        "privacy": "Count Only",
        "notes": "Archived client with no active bridge access.",
    },
    {
        "client_id": "CL-005",
        "display_name": "Bridge Demo Utilities",
        "legal_name": "Bridge Demo Utilities LLC",
        "status": "Active",
        "primary_contact_alias": "bridge.ops.alias",
        "contact_method": "Ops Alias",
        "region": "East",
        "industry": "Utilities",
        "onboarded": "2026-06-01",
        "offboarded": "",
        "ledger_location_alias": "LEDGER-ALIAS-DEMO-005",
        "worker_data_owner": "Client Owned",
        "allowed_worker_count": 10,
        "reported_worker_count": 9,
        "privacy": "Client Controlled",
        "notes": "Active bridge demo client for control-plane checks.",
    },
]


def build_workbook() -> Path:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    wb.properties.title = WORKBOOK_TITLE
    wb.properties.subject = "Tim-only owner control plane"
    wb.properties.creator = "Codex"
    wb.properties.company = "CrewPay"
    wb.properties.description = "Private client control workbook for access, billing, bridge, flags, health, calendar visibility, and support."
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    created = {}
    for sheet_name in SHEET_ORDER:
        created[sheet_name] = wb.create_sheet(title=sheet_name)

    build_instructions_sheet(created["Instructions"])
    build_dashboard_sheet(created["Owner Dashboard"])
    build_client_registry_sheet(created["Client Registry"])
    build_client_access_sheet(created["Client Access Control"])
    build_license_billing_sheet(created["License Billing"])
    build_bridge_registry_sheet(created["Bridge Registry"])
    build_feature_flags_sheet(created["Feature Flags"])
    build_system_health_sheet(created["System Health"])
    build_calendar_visibility_sheet(created["Calendar Visibility"])
    build_support_notes_sheet(created["Support Notes"])
    build_audit_log_sheet(created["Owner Audit Log"])
    build_dropdown_lists_sheet(created["Dropdown Lists"])
    build_data_dictionary_sheet(created["Data Dictionary"])
    build_apps_script_setup_sheet(created["Apps Script Setup"])

    apply_global_data_validations(created)
    apply_conditional_formatting(created)
    set_owner_dashboard_card_layout(created["Owner Dashboard"])
    set_print_and_view_prefs(created)
    apply_tab_colors(created)

    output = Path(__file__).resolve().parent / OUTPUT_FILE
    wb.save(output)
    return output


def build_instructions_sheet(ws):
    title_bar(ws, WORKBOOK_TITLE, WORKBOOK_SUBTITLE, end_col=3)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Instructions"], row=3)
    add_section_header(ws, 4, "What This Workbook Is")
    add_bullet_block(
        ws,
        5,
        [
            "Tim-only private owner control for client status, licensing, bridge health, calendar visibility, feature flags, and support.",
            "Control plane only, not the worker data plane.",
            "Safe for alias-only operational tracking and owner follow-up.",
        ],
    )
    add_section_header(ws, 9, "What This Workbook Is Not")
    add_bullet_block(
        ws,
        10,
        [
            "Not payroll approval or payroll calculation software.",
            "Not the operational CrewPay Ledger workbook.",
            "Not a worker records, proof, or time-entry store.",
        ],
    )
    add_section_header(ws, 14, "Setup Flow")
    add_three_step_flow(
        ws,
        15,
        [
            "Open a Google Sheets copy of the workbook.",
            "Paste the owner-control Apps Script into the workbook-bound project.",
            "Run the self-check before using any owner actions.",
        ],
    )
    add_note_banner(ws, 20, "Privacy boundary: do not store worker names, emails, phone numbers, addresses, proof photos, time entries, or payroll detail here.")
    set_widths(ws, {"A": 18, "B": 94, "C": 26})
    ws.print_area = "A1:C22"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_dashboard_sheet(ws):
    title_bar(ws, WORKBOOK_TITLE, "Formula-driven owner summary cards. No manual entry required.", end_col=12)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    write_headers(ws, TAB_HEADERS["Owner Dashboard"], row=3)
    ws.sheet_view.showGridLines = False
    add_dashboard_group_header(ws, 4, "Client Status", "Live client count and access posture.")
    draw_metric_card(ws, "A5", "C6", "Total Clients", "=COUNTA('Client Registry'!$A$4:$A$200)")
    draw_metric_card(ws, "D5", "F6", "Active Clients", "=COUNTIF('Client Registry'!$D$4:$D$200,\"Active\")")
    draw_metric_card(ws, "G5", "I6", "Paused Clients", "=COUNTIF('Client Registry'!$D$4:$D$200,\"Paused\")")
    draw_metric_card(ws, "J5", "L6", "Suspended Clients", "=COUNTIF('Client Registry'!$D$4:$D$200,\"Suspended\")")

    add_dashboard_group_header(ws, 8, "Billing", "Billing and feature-flag controls.")
    draw_metric_card(ws, "A9", "C10", "Billing Overdue", "=COUNTIF('License Billing'!$D$4:$D$200,\"Overdue\")")
    draw_metric_card(ws, "D9", "F10", "Billing Grace Period", "=COUNTIF('License Billing'!$D$4:$D$200,\"Grace Period\")")
    draw_metric_card(ws, "G9", "I10", "Enabled Feature Flags", "=COUNTIF('Feature Flags'!$D$4:$D$200,\"Enabled\")")
    draw_metric_card(ws, "J9", "L10", "Allowed Worker Count", "=SUM('Client Registry'!$M$4:$M$200)")

    add_dashboard_group_header(ws, 12, "Bridge Health", "Alias-only bridge status and rotation state.")
    draw_metric_card(ws, "A13", "C14", "Bridge Healthy", "=COUNTIF('Bridge Registry'!$C$4:$C$200,\"Healthy\")")
    draw_metric_card(ws, "D13", "F14", "Bridge Warning", "=COUNTIF('Bridge Registry'!$C$4:$C$200,\"Warning\")")
    draw_metric_card(ws, "G13", "I14", "Bridge Failing", "=COUNTIF('Bridge Registry'!$C$4:$C$200,\"Failing\")")
    draw_metric_card(ws, "J13", "L14", "Bridge Disabled", "=COUNTIF('Bridge Registry'!$C$4:$C$200,\"Disabled\")")

    add_dashboard_group_header(ws, 16, "Calendar / Support", "Reference visibility and follow-up queue.")
    draw_metric_card(ws, "A17", "C18", "Calendar Enabled", "=COUNTIF('Calendar Visibility'!$C$4:$C$200,\"Enabled\")")
    draw_metric_card(ws, "D17", "F18", "Open Support Notes", "=COUNTIF('Support Notes'!$F$4:$F$200,\"Open\")+COUNTIF('Support Notes'!$F$4:$F$200,\"Waiting\")")
    draw_metric_card(ws, "G17", "I18", "Open Blockers", "=COUNTIF('System Health'!$J$4:$J$200,\"Blocker\")+COUNTIFS('Support Notes'!$E$4:$E$200,\"Urgent\",'Support Notes'!$F$4:$F$200,\"Open\")+COUNTIFS('Support Notes'!$E$4:$E$200,\"Urgent\",'Support Notes'!$F$4:$F$200,\"Waiting\")")
    draw_metric_card(ws, "J17", "L18", "Paused or Offboarded", "=COUNTIF('Client Registry'!$D$4:$D$200,\"Paused\")+COUNTIF('Client Registry'!$D$4:$D$200,\"Offboarded\")")

    ws.freeze_panes = "A4"
    set_widths(ws, {"A": 16, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16})
    ws.print_area = "A1:L22"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def build_client_registry_sheet(ws):
    title_bar(ws, "Client Registry", "Client-level records only. No worker private data.", end_col=16)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Client Registry"], row=3)
    rows = [
        ["CL-001", "Demo Field Services", "Demo Field Services LLC", "Active", "demo.owner.alias", "Support Alias", "Central", "Field Services", "2026-05-01", "", "LEDGER-ALIAS-DEMO-001", "Client Owned", 14, 12, "Client Controlled", "Primary demo client for owner-control examples."],
        ["CL-002", "Sample Industrial Client", "Sample Industrial Client LLC", "Paused", "sample.billing.alias", "Billing Alias", "West", "Industrial", "2026-05-08", "", "LEDGER-ALIAS-DEMO-002", "Client Owned", 8, 8, "Count Only", "Paused during plan review."],
        ["CL-003", "Training Utility Client", "Training Utility Client Co.", "Suspended", "training.support.alias", "Support Alias", "South", "Utilities", "2026-05-15", "", "LEDGER-ALIAS-DEMO-003", "Client Owned", 20, 19, "Alias Only", "Suspended pending owner review."],
        ["CL-004", "Pilot Construction Client", "Pilot Construction Client LLC", "Offboarded", "pilot.archive.alias", "Registry Alias", "North", "Construction", "2026-04-10", "2026-06-02", "LEDGER-ALIAS-DEMO-004", "Client Exported Summary", 6, 0, "Count Only", "Archived client with no active bridge access."],
        ["CL-005", "Bridge Demo Utilities", "Bridge Demo Utilities LLC", "Active", "bridge.ops.alias", "Ops Alias", "East", "Utilities", "2026-06-01", "", "LEDGER-ALIAS-DEMO-005", "Client Owned", 10, 9, "Client Controlled", "Active bridge demo client."],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"I": "yyyy-mm-dd", "J": "yyyy-mm-dd"})
    format_count_columns(ws, ["M", "N"])
    set_widths(ws, {"A": 12, "B": 24, "C": 24, "D": 16, "E": 20, "F": 16, "G": 14, "H": 16, "I": 14, "J": 14, "K": 24, "L": 20, "M": 18, "N": 24, "O": 18, "P": 34})


def build_client_access_sheet(ws):
    title_bar(ws, "Client Access Control", "Access status for clients only. No worker-level access records.", end_col=13)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Client Access Control"], row=3)
    rows = [
        ["AC-001", "CL-001", "Demo Field Services", "Enabled", "2026-05-01", "", "Initial onboarding", "FALSE", "FALSE", "FALSE", "2026-06-01", "Owner", "Current access with active bridge"],
        ["AC-002", "CL-002", "Sample Industrial Client", "Limited", "2026-05-08", "", "Billing review pending", "TRUE", "FALSE", "TRUE", "2026-06-05", "Owner", "New submissions limited during review"],
        ["AC-003", "CL-003", "Training Utility Client", "Suspended", "2026-05-15", "", "Owner pause requested", "TRUE", "TRUE", "TRUE", "2026-06-08", "Owner", "Bridge and sync disabled"],
        ["AC-004", "CL-004", "Pilot Construction Client", "Disabled", "2026-04-10", "2026-06-02", "Offboarded client", "TRUE", "TRUE", "TRUE", "2026-06-02", "Owner", "All access closed after offboarding"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"E": "yyyy-mm-dd", "F": "yyyy-mm-dd", "K": "yyyy-mm-dd"})
    set_widths(ws, {"A": 16, "B": 12, "C": 24, "D": 14, "E": 16, "F": 16, "G": 22, "H": 18, "I": 14, "J": 20, "K": 16, "L": 16, "M": 28})


def build_license_billing_sheet(ws):
    title_bar(ws, "License Billing", "Client-level billing and allowed-worker counts only.", end_col=12)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["License Billing"], row=3)
    rows = [
        ["LIC-001", "CL-001", "Standard", "Current", "2026-06", "2026-07-01", 14, 12, "INV-ALIAS-001", "PM-ALIAS-001", "", "Current and in good standing"],
        ["LIC-002", "CL-002", "Pro", "Grace Period", "2026-06", "2026-06-30", 8, 8, "INV-ALIAS-002", "PM-ALIAS-002", "2026-06-30", "Grace period under owner review"],
        ["LIC-003", "CL-003", "Custom", "Overdue", "2026-06", "2026-06-25", 20, 19, "INV-ALIAS-003", "PM-ALIAS-003", "2026-06-15", "Billing past due"],
        ["LIC-004", "CL-004", "Demo", "Canceled", "2026-05", "2026-06-02", 6, 0, "INV-ALIAS-004", "PM-ALIAS-004", "", "Offboarded client"],
        ["LIC-005", "CL-005", "Starter", "Trial", "2026-06", "2026-07-10", 10, 9, "INV-ALIAS-005", "PM-ALIAS-005", "", "Trial client during bridge demo"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"E": "yyyy-mm", "F": "yyyy-mm-dd", "K": "yyyy-mm-dd"})
    format_count_columns(ws, ["G", "H"])
    set_widths(ws, {"A": 16, "B": 12, "C": 14, "D": 16, "E": 14, "F": 14, "G": 18, "H": 18, "I": 18, "J": 18, "K": 16, "L": 28})


def build_bridge_registry_sheet(ws):
    title_bar(ws, "Bridge Registry", "Alias-only bridge health and token status. No real endpoint URLs or tokens.", end_col=13)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Bridge Registry"], row=3)
    rows = [
        ["BR-001", "CL-001", "Healthy", "BRIDGE-ALIAS-DEMO-001", "Active", "2026-06-08", "2026-07-08", "2026-06-12", "2026-06-12", "", 0, "", "Stable demo bridge"],
        ["BR-002", "CL-002", "Warning", "BRIDGE-ALIAS-DEMO-002", "Rotate Soon", "2026-06-01", "2026-06-22", "2026-06-11", "2026-06-11", "2026-06-12", 2, "Submission retries detected", "Alias only"],
        ["BR-003", "CL-003", "Failing", "BRIDGE-ALIAS-DEMO-003", "Revoked", "2026-05-30", "2026-06-15", "2026-06-10", "", "2026-06-10", 5, "Token rejected by endpoint", "Requires owner review"],
        ["BR-004", "CL-004", "Disabled", "BRIDGE-ALIAS-DEMO-004", "Not Set", "", "", "2026-06-02", "", "2026-06-02", 0, "Offboarded client", "Bridge disabled"],
        ["BR-005", "CL-005", "Healthy", "BRIDGE-ALIAS-DEMO-005", "Active", "2026-06-05", "2026-07-05", "2026-06-12", "2026-06-12", "", 1, "", "Active demo bridge"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"F": "yyyy-mm-dd", "G": "yyyy-mm-dd", "H": "yyyy-mm-dd", "I": "yyyy-mm-dd", "J": "yyyy-mm-dd"})
    format_count_columns(ws, ["K"])
    set_widths(ws, {"A": 16, "B": 12, "C": 14, "D": 24, "E": 16, "F": 14, "G": 14, "H": 14, "I": 16, "J": 14, "K": 18, "L": 24, "M": 24})


def build_feature_flags_sheet(ws):
    title_bar(ws, "Feature Flags", "Client-level feature toggles only.", end_col=9)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Feature Flags"], row=3)
    rows = [
        ["FF-001", "CL-001", "Owner Dashboard", "Enabled", "2026-06-01", "", "TRUE", "TRUE", "Dashboard access enabled"],
        ["FF-002", "CL-001", "Calendar Visibility", "Enabled", "2026-06-01", "", "TRUE", "FALSE", "Calendar reference enabled"],
        ["FF-003", "CL-002", "Bridge Submit", "Trial", "2026-06-01", "2026-06-30", "FALSE", "FALSE", "Trial feature flag"],
        ["FF-004", "CL-003", "Bridge Submit", "Locked", "2026-05-15", "", "FALSE", "FALSE", "Locked while suspended"],
        ["FF-005", "CL-004", "Legacy Reporting", "Disabled", "2026-04-10", "", "FALSE", "FALSE", "Offboarded client"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"E": "yyyy-mm-dd", "F": "yyyy-mm-dd"})
    set_widths(ws, {"A": 14, "B": 12, "C": 22, "D": 14, "E": 14, "F": 14, "G": 22, "H": 18, "I": 26})


def build_system_health_sheet(ws):
    title_bar(ws, "System Health", "Operational status only. No worker private detail.", end_col=14)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["System Health"], row=3)
    rows = [
        ["SH-001", "CL-001", "2026-06-12", "Healthy", "Healthy", "Healthy", "Enabled", "Healthy", "BACKUP-ALIAS-001", "Info", "All systems normal", "", "", "Demo row"],
        ["SH-002", "CL-002", "2026-06-12", "Warning", "Warning", "Healthy", "Limited", "Healthy", "BACKUP-ALIAS-002", "Warning", "Bridge retry threshold reached", "Review bridge routing", "", "Demo row"],
        ["SH-003", "CL-003", "2026-06-12", "Failing", "Failing", "Warning", "Limited", "Warning", "BACKUP-ALIAS-003", "Blocker", "Bridge submit failing", "Restore bridge status", "", "Demo row"],
        ["SH-004", "CL-004", "2026-06-12", "Disabled", "Disabled", "Disabled", "Disabled", "Disabled", "BACKUP-ALIAS-004", "None", "Offboarded client", "", "2026-06-02", "Demo row"],
        ["SH-005", "CL-005", "2026-06-12", "Healthy", "Healthy", "Healthy", "Enabled", "Healthy", "BACKUP-ALIAS-005", "Info", "Backup verified", "", "", "Demo row"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"C": "yyyy-mm-dd", "D": "yyyy-mm-dd", "E": "yyyy-mm-dd", "F": "yyyy-mm-dd", "G": "yyyy-mm-dd", "H": "yyyy-mm-dd", "I": "yyyy-mm-dd", "L": "yyyy-mm-dd", "M": "yyyy-mm-dd"})
    set_widths(ws, {"A": 14, "B": 12, "C": 14, "D": 14, "E": 14, "F": 16, "G": 14, "H": 14, "I": 18, "J": 14, "K": 24, "L": 22, "M": 14, "N": 16})


def build_calendar_visibility_sheet(ws):
    title_bar(ws, "Calendar Visibility", "Planning/reference only. Calendar is not proof or payroll truth.", end_col=10)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Calendar Visibility"], row=3)
    rows = [
        ["CR-001", "CL-001", "Enabled", "Owner to Client", "Alias Only", "Summary Only", "Client Report", "Owner to Client", "2026-06-12", "Calendar reference only"],
        ["CR-002", "CL-002", "Limited", "Owner to Client", "Count Only", "Summary Only", "Client Report", "Owner to Client", "2026-06-11", "Limited reference only"],
        ["CR-003", "CL-003", "Disabled", "None", "None", "None", "None", "None", "2026-06-10", "Disabled during suspension"],
        ["CR-004", "CL-004", "Disabled", "None", "None", "None", "None", "None", "2026-06-02", "Offboarded client"],
        ["CR-005", "CL-005", "Enabled", "Owner to Client", "Alias Only", "Client Approved", "Worker Export", "Two Way Limited", "2026-06-12", "Owner reviewed"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"B": "yyyy-mm-dd", "I": "yyyy-mm-dd"})
    set_widths(ws, {"A": 14, "B": 12, "C": 16, "D": 20, "E": 16, "F": 18, "G": 16, "H": 18, "I": 18, "J": 24})


def build_support_notes_sheet(ws):
    title_bar(ws, "Support Notes", "Owner notes for follow-up, review, and account support only.", end_col=11)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Support Notes"], row=3)
    rows = [
        ["SN-001", "CL-001", "2026-06-08", "Bridge", "High", "Open", "Bridge retry threshold reached", "Review endpoint alias", "2026-06-14", "", "Need owner review"],
        ["SN-002", "CL-002", "2026-06-09", "Billing", "Urgent", "Waiting", "Grace period ending soon", "Confirm payment alias", "2026-06-13", "", "Escalate if not updated"],
        ["SN-003", "CL-003", "2026-06-10", "Access", "Normal", "Resolved", "Access paused per owner", "No further action", "", "2026-06-11", "Resolved note"],
        ["SN-004", "CL-004", "Offboard", "Low", "Archived", "Archived", "Client archived", "No action", "", "2026-06-02", "Archived note"],
        ["SN-005", "CL-005", "2026-06-12", "Calendar", "Urgent", "Open", "Calendar visibility review needed", "Check alias-only rule", "2026-06-15", "", "Calendar rule owner review"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    set_widths(ws, {"A": 14, "B": 12, "C": 14, "D": 14, "E": 12, "F": 12, "G": 26, "H": 22, "I": 16, "J": 14, "K": 24})


def build_audit_log_sheet(ws):
    title_bar(ws, "Owner Audit Log", "Append-only owner activity trail.", end_col=10)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Owner Audit Log"], row=3)
    rows = [
        ["AUD-001", "2026-06-10 09:00", "Owner", "Access", "CL-001", "Enabled", "Limited", "Plan review", "Owner review note"],
        ["AUD-002", "2026-06-10 09:15", "Owner", "Billing", "CL-002", "Current", "Grace Period", "Invoice follow-up", "Owner review note"],
        ["AUD-003", "2026-06-10 09:30", "Owner", "Bridge", "CL-003", "Warning", "Failing", "Bridge triage", "Owner review note"],
        ["AUD-004", "2026-06-10 09:45", "Owner", "Support", "CL-005", "Open", "Waiting", "Support ticket created", "Owner review note"],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    format_date_columns(ws, {"B": "yyyy-mm-dd", "C": "yyyy-mm-dd", "I": "yyyy-mm-dd"})
    set_widths(ws, {"A": 14, "B": 20, "C": 14, "D": 14, "E": 12, "F": 16, "G": 16, "H": 18, "I": 16, "J": 24})


def build_dropdown_lists_sheet(ws):
    title_bar(ws, "Dropdown Lists", "Reference values used by data validation on control-plane tabs.", end_col=len(TAB_HEADERS["Dropdown Lists"]))
    ws.freeze_panes = "A4"
    headers = TAB_HEADERS["Dropdown Lists"]
    write_headers(ws, headers, row=3)

    dropdowns = {
        "Client Status": ["Prospect", "Active", "Paused", "Suspended", "Offboarded"],
        "Access Status": ["Enabled", "Limited", "Suspended", "Disabled"],
        "Billing Status": ["Trial", "Current", "Grace Period", "Overdue", "Canceled"],
        "Plan Tier": ["Demo", "Starter", "Standard", "Pro", "Custom"],
        "Bridge Status": ["Not Configured", "Healthy", "Warning", "Failing", "Disabled"],
        "Token Status": ["Not Set", "Active", "Rotate Soon", "Expired", "Revoked"],
        "Feature Status": ["Enabled", "Disabled", "Trial", "Locked"],
        "Calendar Visibility Status": ["Disabled", "Limited", "Enabled"],
        "Calendar Status": ["Disabled", "Limited", "Enabled"],
        "Worker Detail Exposure": ["None", "Count Only", "Alias Only"],
        "Schedule Detail Exposure": ["None", "Summary Only", "Client Approved"],
        "Proof Source": ["Client Ledger", "Worker Export", "Client Report", "None"],
        "Sync Direction": ["None", "Owner to Client", "Client to Owner", "Two Way Limited"],
        "Priority": ["Low", "Normal", "High", "Urgent"],
        "Support Status": ["Open", "Waiting", "Resolved", "Archived"],
        "Issue Severity": ["None", "Info", "Warning", "Blocker"],
        "Boolean": ["TRUE", "FALSE"],
        "Worker Data Owner": ["Client Owned", "Client Exported Summary", "Owner Count Only"],
        "Data Privacy Level": ["Count Only", "Alias Only", "Client Controlled"],
        "Ledger Status": ["Healthy", "Warning", "Failing", "Disabled"],
        "Worker App Status": ["Healthy", "Warning", "Failing", "Disabled"],
        "Backup Status": ["Healthy", "Warning", "Missing", "Disabled"],
    }

    for column_index, header in enumerate(headers, start=1):
        values = dropdowns[header]
        ws.cell(4, column_index, header)
        for row_index, value in enumerate(values, start=5):
            ws.cell(row_index, column_index, value)
        ws.column_dimensions[get_column_letter(column_index)].width = max(len(header) + 4, 18)
        header_style_cells(ws, 4, column_index, column_index, fill=THEME["blue_2"], font_color=THEME["navy"])
        style_rows(ws, 5, 4 + len(values), column_index, column_index)

    set_widths(ws, {get_column_letter(i): max(len(header) + 4, 18) for i, header in enumerate(headers, start=1)})


def build_data_dictionary_sheet(ws):
    title_bar(ws, "Data Dictionary", "Major fields and privacy boundaries for each owner-control tab.", end_col=4)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Data Dictionary"], row=3)
    rows = [
        ["Instructions", "Workbook purpose, privacy boundary, control-plane rules, and placeholder-only bridge guidance.", "Control plane only.", "No worker private data."],
        ["Owner Dashboard", "Formula-driven summary cards and operating notes.", "Control plane only.", "No manual inputs required."],
        ["Client Registry", "Client ID, display name, legal name, status, allowed counts, privacy level, notes.", "Client-level only.", "No worker identity or contact data."],
        ["Client Access Control", "Access status, start/end dates, access reason, disabled flags, review metadata.", "Client-level control.", "No worker records."],
        ["License Billing", "Plan tier, billing status, billing period, renewal, allowed count, billing count.", "Client-level billing.", "No worker payroll detail."],
        ["Bridge Registry", "Alias-only bridge health, token status, rotation timing, error summary, counts.", "No real endpoint URLs or tokens.", "Alias/status fields only."],
        ["Feature Flags", "Client-specific feature controls and effective dates.", "Client-level control.", "No worker detail."],
        ["System Health", "Ledger, bridge, worker app, calendar, backup, issue severity, and owner action fields.", "System-level only.", "No worker data."],
        ["Calendar Visibility", "Calendar use, exposure levels, proof source, and sync direction.", "Planning/reference only.", "Calendar never becomes payroll truth."],
        ["Support Notes", "Support triage, priorities, actions, and follow-up dates.", "Owner support only.", "No worker private notes."],
        ["Owner Audit Log", "Append-only history of owner actions and value changes.", "Audit trail only.", "Use aliases, not emails."],
        ["Dropdown Lists", "Reference values for status fields and booleans.", "Helper sheet only.", "No private data."],
        ["Apps Script Setup", "Manual setup steps for the workbook-bound owner-control script.", "Setup instructions only.", "No worker data connections."],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    add_forbidden_data_section(ws, start_row=20)
    set_widths(ws, {"A": 20, "B": 68, "C": 28, "D": 28})


def build_apps_script_setup_sheet(ws):
    title_bar(ws, "Apps Script Setup", "Paste the separate owner-control script into a Google Sheets copy of this workbook only.", end_col=3)
    ws.freeze_panes = "A4"
    write_headers(ws, TAB_HEADERS["Apps Script Setup"], row=3)
    rows = [
        ["1", "Open the workbook copy in Google Sheets, then open Extensions > Apps Script.", "Workbook-bound setup only."],
        ["2", "Paste the contents of apps_script/CrewPay_Owner_Control.gs into a separate script file named CrewPay_Owner_Control.gs.", "Do not paste the ledger helper scripts here."],
        ["3", "Save the script project, reload the workbook, and use the CrewPay Owner Control menu.", "Menu functions use aliases and placeholders only."],
        ["4", "Run Refresh Owner Dashboard and the logging actions on sample rows only.", "No worker data connections."],
        ["5", "Confirm the script never uses GmailApp, CalendarApp, DriveApp, UrlFetchApp, ContentService, doGet, or doPost.", "No external API calls."],
        ["6", "Keep the operational CrewPay Ledger workbook and its helper scripts separate from this private owner workbook.", "Separate control plane."],
    ]
    write_table(ws, rows, start_row=4, start_col=1)
    set_widths(ws, {"A": 10, "B": 92, "C": 28})


def apply_global_data_validations(sheets):
    add_list_validation(sheets["Client Registry"], "D4:D200", "'Dropdown Lists'!$A$5:$A$9")
    add_list_validation(sheets["Client Registry"], "L4:L200", "'Dropdown Lists'!$Q$5:$Q$7")
    add_list_validation(sheets["Client Registry"], "O4:O200", "'Dropdown Lists'!$R$5:$R$7")
    add_whole_number_validation(sheets["Client Registry"], "M4:N200", 0)

    add_list_validation(sheets["Client Access Control"], "D4:D200", "'Dropdown Lists'!$B$5:$B$8")
    add_boolean_validation(sheets["Client Access Control"], "H4:J200")

    add_list_validation(sheets["License Billing"], "C4:C200", "'Dropdown Lists'!$D$5:$D$9")
    add_list_validation(sheets["License Billing"], "D4:D200", "'Dropdown Lists'!$C$5:$C$9")
    add_whole_number_validation(sheets["License Billing"], "G4:H200", 0)

    add_list_validation(sheets["Bridge Registry"], "C4:C200", "'Dropdown Lists'!$E$5:$E$9")
    add_list_validation(sheets["Bridge Registry"], "E4:E200", "'Dropdown Lists'!$F$5:$F$9")
    add_whole_number_validation(sheets["Bridge Registry"], "K4:K200", 0)

    add_list_validation(sheets["Feature Flags"], "D4:D200", "'Dropdown Lists'!$G$5:$G$8")
    add_boolean_validation(sheets["Feature Flags"], "G4:H200")

    add_list_validation(sheets["System Health"], "D4:D200", "'Dropdown Lists'!$S$5:$S$8")
    add_list_validation(sheets["System Health"], "E4:E200", "'Dropdown Lists'!$E$5:$E$9")
    add_list_validation(sheets["System Health"], "F4:F200", "'Dropdown Lists'!$T$5:$T$8")
    add_list_validation(sheets["System Health"], "G4:G200", "'Dropdown Lists'!$V$5:$V$7")
    add_list_validation(sheets["System Health"], "H4:H200", "'Dropdown Lists'!$U$5:$U$8")
    add_list_validation(sheets["System Health"], "J4:J200", "'Dropdown Lists'!$O$5:$O$8")

    add_list_validation(sheets["Calendar Visibility"], "C4:C200", "'Dropdown Lists'!$H$5:$H$7")
    add_list_validation(sheets["Calendar Visibility"], "E4:E200", "'Dropdown Lists'!$I$5:$I$7")
    add_list_validation(sheets["Calendar Visibility"], "F4:F200", "'Dropdown Lists'!$J$5:$J$7")
    add_list_validation(sheets["Calendar Visibility"], "G4:G200", "'Dropdown Lists'!$K$5:$K$8")
    add_list_validation(sheets["Calendar Visibility"], "H4:H200", "'Dropdown Lists'!$L$5:$L$8")

    add_list_validation(sheets["Support Notes"], "E4:E200", "'Dropdown Lists'!$M$5:$M$8")
    add_list_validation(sheets["Support Notes"], "F4:F200", "'Dropdown Lists'!$N$5:$N$8")


def apply_conditional_formatting(sheets):
    status_rules = {
        "Client Registry": {
            "D": {
                "Active": THEME["green"],
                "Paused": THEME["yellow"],
                "Suspended": THEME["orange"],
                "Offboarded": THEME["gray"],
                "Prospect": THEME["blue_2"],
            },
        },
        "Client Access Control": {
            "D": {
                "Enabled": THEME["green"],
                "Limited": THEME["yellow"],
                "Suspended": THEME["orange"],
                "Disabled": THEME["gray"],
            },
        },
        "License Billing": {
            "D": {
                "Current": THEME["green"],
                "Grace Period": THEME["yellow"],
                "Overdue": THEME["red"],
                "Canceled": THEME["gray"],
                "Trial": THEME["blue_2"],
            },
            "C": {
                "Demo": THEME["blue_2"],
                "Starter": THEME["blue"],
                "Standard": THEME["green"],
                "Pro": THEME["teal"],
                "Custom": THEME["orange"],
            },
        },
        "Bridge Registry": {
            "C": {
                "Healthy": THEME["green"],
                "Warning": THEME["yellow"],
                "Failing": THEME["red"],
                "Disabled": THEME["gray"],
                "Not Configured": THEME["blue_2"],
            },
            "E": {
                "Active": THEME["green"],
                "Rotate Soon": THEME["yellow"],
                "Expired": THEME["orange"],
                "Revoked": THEME["red"],
                "Not Set": THEME["gray"],
            },
        },
        "Feature Flags": {
            "D": {
                "Enabled": THEME["green"],
                "Disabled": THEME["gray"],
                "Trial": THEME["yellow"],
                "Locked": THEME["orange"],
            },
        },
        "System Health": {
            "D": {"Healthy": THEME["green"], "Warning": THEME["yellow"], "Failing": THEME["red"], "Disabled": THEME["gray"]},
            "E": {"Healthy": THEME["green"], "Warning": THEME["yellow"], "Failing": THEME["red"], "Disabled": THEME["gray"]},
            "F": {"Healthy": THEME["green"], "Warning": THEME["yellow"], "Failing": THEME["red"], "Disabled": THEME["gray"]},
            "G": {"Enabled": THEME["green"], "Limited": THEME["yellow"], "Disabled": THEME["gray"]},
            "H": {"Healthy": THEME["green"], "Warning": THEME["yellow"], "Missing": THEME["orange"], "Disabled": THEME["gray"]},
            "J": {"None": THEME["gray_2"], "Info": THEME["blue_2"], "Warning": THEME["yellow"], "Blocker": THEME["red"]},
        },
        "Calendar Visibility": {
            "C": {"Disabled": THEME["gray"], "Limited": THEME["yellow"], "Enabled": THEME["green"]},
        },
        "Support Notes": {
            "E": {"Low": THEME["green"], "Normal": THEME["blue_2"], "High": THEME["yellow"], "Urgent": THEME["red"]},
            "F": {"Open": THEME["yellow"], "Waiting": THEME["orange"], "Resolved": THEME["green"], "Archived": THEME["gray"]},
        },
    }

    for sheet_name, column_rules in status_rules.items():
        ws = sheets[sheet_name]
        for col_letter, mapping in column_rules.items():
            for status_value, fill_color in mapping.items():
                formula = f'${col_letter}4="{status_value}"'
                ws.conditional_formatting.add(
                    f"{col_letter}4:{col_letter}200",
                    FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=fill_color)),
                )

    dashboard = sheets["Owner Dashboard"]
    for cell_range in ["A4:C5", "D4:F5", "G4:I5", "J4:L5", "A7:C8", "D7:F8", "G7:I8", "J7:L8", "A10:C11", "D10:F11", "G10:I11", "J10:L11", "A13:C14"]:
        dashboard.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=["TRUE"], fill=PatternFill("solid", fgColor=THEME["white"]))
        )


def set_owner_dashboard_card_layout(ws):
    for row in range(4, 23):
        ws.row_dimensions[row].height = 22
    for row in [4, 8, 12, 16]:
        ws.row_dimensions[row].height = 20
    for row in [5, 9, 13, 17]:
        ws.row_dimensions[row + 1].height = 30
    set_widths(ws, {c: 14 for c in [get_column_letter(i) for i in range(1, 13)]})
    ws["A20"] = "Operating Notes"
    ws["A20"].font = Font(name="Aptos", bold=True, color=THEME["navy"])
    ws["B20"] = "Formula counts update from the control tabs."
    ws["A21"] = "Privacy Note"
    ws["A21"].font = Font(name="Aptos", bold=True, color=THEME["navy"])
    ws["B21"] = "Control plane only — no worker private records."
    ws["A22"] = "Operating Note"
    ws["B22"] = "Use aliases and count-level fields only."
    style_rows(ws, 20, 22, 1, 2)
    ws.print_area = "A1:L22"


def set_print_and_view_prefs(sheets):
    for sheet_name, ws in sheets.items():
        if sheet_name in {"Owner Dashboard", "Instructions", "Data Dictionary", "Apps Script Setup"}:
            ws.sheet_view.showGridLines = False
        if sheet_name not in {"Instructions", "Owner Dashboard", "Data Dictionary", "Apps Script Setup"}:
            ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        else:
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0


def apply_tab_colors(sheets):
    colors = {
        "Instructions": "4B6584",
        "Owner Dashboard": "17324D",
        "Client Registry": "2F6F68",
        "Client Access Control": "5B6C9A",
        "License Billing": "8B5E3C",
        "Bridge Registry": "2D5B7E",
        "Feature Flags": "7C5E9B",
        "System Health": "556B2F",
        "Calendar Visibility": "4F6D7A",
        "Support Notes": "7A4B4B",
        "Owner Audit Log": "4D4D4D",
        "Dropdown Lists": "607D8B",
        "Data Dictionary": "3F51B5",
        "Apps Script Setup": "2E7D32",
    }
    for sheet_name, color in colors.items():
        sheets[sheet_name].sheet_properties.tabColor = color


def write_headers(ws, headers, row=3, start_col=1):
    for index, header in enumerate(headers, start=start_col):
        ws.cell(row, index, header)
    header_style_cells(ws, row, start_col, start_col + len(headers) - 1)


def header_style_cells(ws, row, start_col, end_col, fill=THEME["navy"], font_color=THEME["white"]):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=THEME["navy"]))


def style_rows(ws, start_row, end_row, start_col, end_col):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=THEME["gray_2"])


def write_table(ws, rows, start_row, start_col):
    for row_offset, row_values in enumerate(rows, start=0):
        for col_offset, value in enumerate(row_values, start=0):
            cell = ws.cell(start_row + row_offset, start_col + col_offset, value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=THEME["gray_2"])
    style_rows(ws, start_row, start_row + len(rows) - 1, start_col, start_col + len(rows[0]) - 1)


def draw_metric_card(ws, start_cell, end_cell, label, formula):
    start_col = column_index(start_cell)
    start_row = row_index(start_cell)
    end_col = column_index(end_cell)
    end_row = row_index(end_cell)

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    ws.cell(start_row, start_col, label)
    ws.cell(start_row, start_col).font = Font(name="Aptos", bold=True, color=THEME["navy"])
    ws.cell(start_row, start_col).alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
    ws.cell(start_row + 1, start_col, formula)
    ws.cell(start_row + 1, start_col).font = Font(name="Aptos", bold=True, size=18, color=THEME["navy"])
    ws.cell(start_row + 1, start_col).alignment = Alignment(horizontal="center", vertical="center")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=THEME["white"])
            ws.cell(row, col).border = THIN_BORDER
    ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor=THEME["blue_2"])
    ws.cell(start_row + 1, start_col).fill = PatternFill("solid", fgColor=THEME["white"])


def add_note_banner(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row, 1, text)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=THEME["blue_2"])
    ws.cell(row, 1).font = Font(name="Aptos", italic=True, color=THEME["muted"])
    ws.cell(row, 1).alignment = Alignment(wrap_text=True)
    for col in range(1, 4):
        ws.cell(row, col).border = THIN_BORDER


def add_forbidden_data_section(ws, start_row):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    ws.cell(start_row, 1, "Forbidden Data")
    ws.cell(start_row, 1).font = Font(name="Aptos", bold=True, color=THEME["navy"])
    ws.cell(start_row, 1).fill = PatternFill("solid", fgColor=THEME["orange"])
    forbidden = [
        "Worker Email",
        "Worker Phone",
        "Worker Address",
        "Worker SSN",
        "Time Entry",
        "Hours Worked",
        "Pay Rate",
        "Gross Pay",
        "Net Pay",
        "Payroll Detail",
        "Proof Photo",
        "Private Worker Notes",
        "Worker Notes",
        "Worker Time Detail",
        "Worker Payroll Detail",
    ]
    for offset, value in enumerate(forbidden, start=1):
        ws.cell(start_row + offset, 1, value)
        ws.cell(start_row + offset, 2, "Must never be stored in this workbook.")
        ws.cell(start_row + offset, 3, "Control plane only.")
        style_rows(ws, start_row + offset, start_row + offset, 1, 3)


def add_section_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row, 1, text)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=THEME["navy"])
    ws.cell(row, 1).font = Font(name="Aptos", bold=True, color=THEME["white"])
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    for col in range(1, 4):
        ws.cell(row, col).border = THIN_BORDER


def add_three_step_flow(ws, row, steps):
    for index, text in enumerate(steps, start=0):
        current_row = row + index
        ws.cell(current_row, 1, f"Step {index + 1}")
        ws.cell(current_row, 2, text)
        ws.cell(current_row, 3, "Owner setup")
        style_rows(ws, current_row, current_row, 1, 3)
        ws.cell(current_row, 1).fill = PatternFill("solid", fgColor=THEME["blue_2"])
        ws.cell(current_row, 1).font = Font(name="Aptos", bold=True, color=THEME["navy"])


def add_bullet_block(ws, row, bullets):
    for index, text in enumerate(bullets, start=0):
        current_row = row + index
        ws.cell(current_row, 1, "•")
        ws.cell(current_row, 2, text)
        ws.cell(current_row, 3, "Owner guidance")
        style_rows(ws, current_row, current_row, 1, 3)
        ws.cell(current_row, 1).font = Font(name="Aptos", bold=True, color=THEME["navy"])


def add_dashboard_group_header(ws, row, title, subtitle):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row, 1, f"{title} — {subtitle}")
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=THEME["navy"])
    ws.cell(row, 1).font = Font(name="Aptos", bold=True, color=THEME["white"])
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    for col in range(1, 13):
        ws.cell(row, col).border = THIN_BORDER


def add_dashboard_notes(ws):
    ws.merge_cells(start_row=20, start_column=1, end_row=20, end_column=12)
    ws.cell(20, 1, "Control plane only — no worker private records.")
    ws.cell(20, 1).fill = PatternFill("solid", fgColor=THEME["orange"])
    ws.cell(20, 1).font = Font(name="Aptos", bold=True, color=THEME["navy"])
    ws.cell(20, 1).alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=21, start_column=1, end_row=21, end_column=12)
    ws.cell(21, 1, "Review status counts before changing access, billing, bridge, calendar, or support settings.")
    ws.cell(21, 1).fill = PatternFill("solid", fgColor=THEME["blue_2"])
    ws.cell(21, 1).font = Font(name="Aptos", italic=True, color=THEME["muted"])
    ws.cell(21, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=22, start_column=1, end_row=22, end_column=12)
    ws.cell(22, 1, "Use the self-check menu first if the workbook was copied or renamed.")
    ws.cell(22, 1).fill = PatternFill("solid", fgColor=THEME["gray_2"])
    ws.cell(22, 1).font = Font(name="Aptos", color=THEME["text"])
    ws.cell(22, 1).alignment = Alignment(horizontal="left", vertical="center")

    for row in [20, 21, 22]:
        for col in range(1, 13):
            ws.cell(row, col).border = THIN_BORDER


def format_date_columns(ws, columns_with_formats):
    for col_letter, number_format in columns_with_formats.items():
        for row in range(4, ws.max_row + 1):
            ws[f"{col_letter}{row}"].number_format = number_format


def format_count_columns(ws, columns):
    for col_letter in columns:
        for row in range(4, ws.max_row + 1):
            ws[f"{col_letter}{row}"].number_format = "0"


def add_list_validation(ws, cell_range, formula_range):
    validation = DataValidation(type="list", formula1=f"={formula_range}", allow_blank=True)
    validation.promptTitle = "Choose from list"
    validation.prompt = "Use the allowed private-control values only."
    validation.errorTitle = "Invalid value"
    validation.error = "Select a value from the dropdown list."
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_boolean_validation(ws, cell_range):
    validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    validation.promptTitle = "Boolean"
    validation.prompt = "Use TRUE or FALSE."
    validation.errorTitle = "Invalid value"
    validation.error = "Choose TRUE or FALSE."
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_whole_number_validation(ws, cell_range, min_value=0):
    validation = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=str(min_value), allow_blank=True)
    validation.promptTitle = "Count"
    validation.prompt = "Enter a count value."
    validation.errorTitle = "Invalid count"
    validation.error = "Enter a whole number greater than or equal to the minimum."
    ws.add_data_validation(validation)
    validation.add(cell_range)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def title_bar(ws, title, subtitle=None, end_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=THEME["navy"])
    ws["A1"].font = Font(name="Aptos", bold=True, size=18, color=THEME["white"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws["A2"] = subtitle
        ws["A2"].fill = PatternFill("solid", fgColor=THEME["blue_2"])
        ws["A2"].font = Font(name="Aptos", italic=True, color=THEME["muted"])
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 24
    for row in [1, 2]:
        for col in range(1, end_col + 1):
            ws.cell(row, col).border = THIN_BORDER


def column_index(cell_ref):
    letters = "".join([c for c in cell_ref if c.isalpha()])
    result = 0
    for char in letters:
        result = result * 26 + (ord(char.upper()) - ord("A") + 1)
    return result


def row_index(cell_ref):
    digits = "".join([c for c in cell_ref if c.isdigit()])
    return int(digits)


def main():
    output = build_workbook()
    print(output)


if __name__ == "__main__":
    main()
