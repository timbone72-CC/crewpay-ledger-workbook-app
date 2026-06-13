from __future__ import annotations

import re
import subprocess
import sys
from pathlib import Path

LOCAL_DEPS = Path(__file__).resolve().parent / ".deps" / "usr" / "lib" / "python3" / "dist-packages"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_FILE = ROOT / "CrewPay_Owner_Control_Workbook.xlsx"
SCRIPT_FILE = ROOT / "apps_script" / "CrewPay_Owner_Control.gs"
LEGACY_WORKBOOK_FILE = ROOT / "CrewPay_Ledger_Workbook.xlsx"
LEGACY_LEDGER_SCRIPT = ROOT / "apps_script" / "Code.gs"
LEGACY_ORIGINAL_FINAL_SCRIPT = ROOT / "apps_script" / "CrewPay_Ledger_ORIGINAL_FINAL.gs"
LEGACY_BRIDGE_SCRIPT = ROOT / "apps_script" / "CrewPay_Ledger_BRIDGE.gs"

EXPECTED_SHEETS = [
    "Instructions",
    "Owner Dashboard",
    "Client Registry",
    "Client Access Control",
    "License Billing",
    "Bridge Registry",
    "Feature Flags",
    "System Health",
    "Calendar Visibility",
    "Support Notes",
    "Owner Audit Log",
    "Dropdown Lists",
    "Data Dictionary",
    "Apps Script Setup",
]

EXPECTED_HEADERS = {
    "Instructions": ["Section", "Details", "Privacy Boundary"],
    "Owner Dashboard": ["Operating Note", "Details"],
    "Client Registry": [
        "Client ID",
        "Client Display Name",
        "Client Legal Name",
        "Client Status",
        "Primary Contact Alias",
        "Contact Method",
        "Region",
        "Industry",
        "Onboarded Date",
        "Offboarded Date",
        "Client Ledger Location Alias",
        "Worker Data Owner",
        "Allowed Worker Count",
        "Current Worker Count Reported",
        "Data Privacy Level",
        "Notes",
    ],
    "Client Access Control": [
        "Access Record ID",
        "Client ID",
        "Client Display Name",
        "Access Status",
        "Access Start Date",
        "Access End Date",
        "Access Reason",
        "Disabled New Submissions",
        "Disabled Bridge",
        "Disabled Calendar Sync",
        "Last Access Review",
        "Reviewed By Alias",
        "Notes",
    ],
    "License Billing": [
        "License Record ID",
        "Client ID",
        "Plan Tier",
        "Billing Status",
        "Billing Period",
        "Renewal Date",
        "Allowed Worker Count",
        "Billing Worker Count",
        "Last Invoice Alias",
        "Payment Method Alias",
        "Grace Period Ends",
        "Billing Notes",
    ],
    "Bridge Registry": [
        "Bridge Record ID",
        "Client ID",
        "Bridge Status",
        "Bridge Endpoint Alias",
        "Token Status",
        "Token Last Rotated",
        "Token Rotation Due",
        "Last Health Check",
        "Last Successful Submit",
        "Last Failed Submit",
        "Pending Intake Count",
        "Last Error Summary",
        "Notes",
    ],
    "Feature Flags": [
        "Flag Record ID",
        "Client ID",
        "Feature Name",
        "Feature Status",
        "Effective Date",
        "Expiration Date",
        "Requires Billing Good Standing",
        "Requires Bridge Healthy",
        "Notes",
    ],
    "System Health": [
        "Health Record ID",
        "Client ID",
        "Check Date",
        "Ledger Status",
        "Bridge Status",
        "Worker App Status",
        "Calendar Status",
        "Backup Status",
        "Last Backup Alias",
        "Issue Severity",
        "Issue Summary",
        "Owner Action Needed",
        "Resolved Date",
        "Notes",
    ],
    "Calendar Visibility": [
        "Calendar Rule ID",
        "Client ID",
        "Calendar Visibility Status",
        "Allowed Calendar Use",
        "Worker Detail Exposure",
        "Schedule Detail Exposure",
        "Proof Source",
        "Sync Direction",
        "Last Calendar Sync",
        "Notes",
    ],
    "Support Notes": [
        "Support Note ID",
        "Client ID",
        "Note Date",
        "Note Type",
        "Priority",
        "Status",
        "Summary",
        "Owner Next Action",
        "Follow-up Date",
        "Resolved Date",
        "Notes",
    ],
    "Owner Audit Log": [
        "Audit ID",
        "Timestamp",
        "Actor Alias",
        "Area",
        "Client ID",
        "Action",
        "Previous Value",
        "New Value",
        "Reason",
        "Notes",
    ],
    "Dropdown Lists": [
        "Client Status",
        "Access Status",
        "Billing Status",
        "Plan Tier",
        "Bridge Status",
        "Token Status",
        "Feature Status",
        "Calendar Visibility Status",
        "Calendar Status",
        "Worker Detail Exposure",
        "Schedule Detail Exposure",
        "Proof Source",
        "Sync Direction",
        "Priority",
        "Support Status",
        "Issue Severity",
        "Boolean",
        "Worker Data Owner",
        "Data Privacy Level",
        "Ledger Status",
        "Worker App Status",
        "Backup Status",
    ],
    "Data Dictionary": ["Tab", "Major Fields", "Privacy Boundary", "Notes"],
    "Apps Script Setup": ["Step", "Instructions", "Notes"],
}

DASHBOARD_CARD_CELLS = {
    "A5": "A6",
    "D5": "D6",
    "G5": "G6",
    "J5": "J6",
    "A9": "A10",
    "D9": "D10",
    "G9": "G10",
    "J9": "J10",
    "A13": "A14",
    "D13": "D14",
    "G13": "G14",
    "J13": "J14",
    "A17": "A18",
}

FORBIDDEN_COLUMN_TERMS = [
    "Worker Email",
    "Worker Phone",
    "Worker Address",
    "Worker SSN",
    "Time Entry",
    "Hours Worked",
    "Pay Rate",
    "Gross Pay",
    "Net Pay",
    "Payroll Detail",
    "Proof Photo",
    "Private Worker Notes",
    "Worker Notes",
    "Worker Time Detail",
    "Worker Payroll Detail",
]

FORBIDDEN_CELL_PATTERNS = [
    re.compile(r"https?://", re.I),
    re.compile(r"script\.google\.com", re.I),
    re.compile(r"docs\.google\.com", re.I),
    re.compile(r"AKfy[a-zA-Z0-9_-]+", re.I),
    re.compile(r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b"),
    re.compile(r"\b\d{3}[-.]\d{3}[-.]\d{4}\b"),
    re.compile(r"\(\d{3}\)\s*\d{3}[-.]\d{4}"),
]


def fail(message: str) -> None:
    raise AssertionError(message)


def check_file(path: Path) -> None:
    if not path.exists():
        fail(f"Missing required file: {path}")


def check_git_path_clean(path: str) -> None:
    diff = subprocess.run(
        ["git", "diff", "--quiet", "--", path],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )
    if diff.returncode != 0:
        fail(f"Protected path has uncommitted changes: {path}")

    untracked = subprocess.run(
        ["git", "ls-files", "--others", "--exclude-standard", "--", path],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    if untracked.stdout.strip():
        fail(f"Protected path has untracked changes: {path}")


def check_sheet_headers(ws, expected_headers):
    headers = [ws.cell(3, col).value for col in range(1, len(expected_headers) + 1)]
    if headers != expected_headers:
        fail(f"{ws.title} headers mismatch: {headers!r}")


def check_dashboard_formulas(ws):
    for start_cell, value_cell in DASHBOARD_CARD_CELLS.items():
        value = ws[value_cell].value
        if not (isinstance(value, str) and value.startswith("=")):
            fail(f"Dashboard metric cell {value_cell} does not contain a formula.")


def check_data_validations(wb):
    required_ranges = {
        "Client Registry": ["D4:D200", "L4:L200", "O4:O200", "M4:N200"],
        "Client Access Control": ["D4:D200", "H4:J200"],
        "License Billing": ["C4:C200", "D4:D200", "G4:H200"],
        "Bridge Registry": ["C4:C200", "E4:E200", "K4:K200"],
        "Feature Flags": ["D4:D200", "G4:H200"],
        "System Health": ["D4:D200", "E4:E200", "F4:F200", "G4:G200", "H4:H200", "J4:J200"],
        "Calendar Visibility": ["C4:C200", "E4:E200", "F4:F200", "G4:G200", "H4:H200"],
        "Support Notes": ["E4:E200", "F4:F200"],
    }

    for sheet_name, expected_ranges in required_ranges.items():
        ws = wb[sheet_name]
        sqref_text = " ".join(str(dv.sqref) for dv in ws.data_validations.dataValidation)
        for expected in expected_ranges:
            if expected not in sqref_text:
                fail(f"{sheet_name} missing validation for range {expected}")


def check_forbidden_terms_in_columns(wb):
    for ws in wb.worksheets:
        if ws.title == "Data Dictionary":
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() in FORBIDDEN_COLUMN_TERMS:
                    fail(f"Forbidden owner workbook term appears as a cell value: {ws.title}!{cell.coordinate}={value}")


def check_forbidden_patterns(wb):
    for ws in wb.worksheets:
        if ws.title == "Data Dictionary":
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                for pattern in FORBIDDEN_CELL_PATTERNS:
                    if pattern.search(value):
                        fail(f"Forbidden private or URL-like text found at {ws.title}!{cell.coordinate}: {value}")
                if any(
                    term in value
                    for term in [
                        "Worker Email",
                        "Worker Phone",
                        "Worker Address",
                        "Worker SSN",
                        "Payroll Detail",
                        "Time Entry",
                        "Proof Photo",
                        "Private Worker Notes",
                        "Worker Notes",
                        "Worker Time Detail",
                        "Worker Payroll Detail",
                    ]
                ):
                    fail(f"Forbidden worker-level data text found at {ws.title}!{cell.coordinate}: {value}")


def check_script() -> None:
    check_file(SCRIPT_FILE)
    text = SCRIPT_FILE.read_text(encoding="utf-8")
    required_functions = [
        "function onOpen()",
        "function runOwnerControlSelfCheck()",
        "function refreshOwnerDashboard()",
        "function logClientAccessChange()",
        "function logBillingStatusChange()",
        "function logBridgeHealthCheck()",
        "function createSupportNote()",
        "function aboutCrewPayOwnerControl()",
        "function validateOwnerControlWorkbookStructure_(",
        "function appendByHeaders_(",
    ]
    for marker in required_functions:
        if marker not in text:
            fail(f"Missing required Apps Script function: {marker}")
    required_guard_markers = [
        "OWNER_CONTROL.FORBIDDEN_TABS",
        "OWNER_CONTROL.WRITE_TARGETS",
        "EXPECTED_HEADERS_BY_SHEET_",
        "Missing or mismatched headers",
        "This workbook looks like the operational ledger workbook",
    ]
    for marker in required_guard_markers:
        if marker not in text:
            fail(f"Missing required guard logic in Apps Script: {marker}")
    forbidden = [
        "GmailApp",
        "CalendarApp",
        "DriveApp",
        "UrlFetchApp",
        "ContentService",
        "doGet",
        "doPost",
        "CP_BRIDGE_TOKEN",
        "Session.getActiveUser().getEmail",
        "external fetch",
        "token-setting logic",
        "http://",
        "https://",
        "Worker Email",
        "Worker Phone",
        "Worker Address",
        "Worker SSN",
        "Time Entry",
        "Hours Worked",
        "Pay Rate",
        "Gross Pay",
        "Net Pay",
        "Payroll Detail",
        "Proof Photo",
        "Private Worker Notes",
        "Worker Time Detail",
        "Worker Payroll Detail",
        "pending time-entry submit",
        "worker proof export",
        "worker email lookup",
    ]
    for term in forbidden:
        if term in text:
            fail(f"Forbidden Apps Script term present: {term}")


def check_legacy_scripts_unchanged() -> None:
    for path in [LEGACY_LEDGER_SCRIPT, LEGACY_ORIGINAL_FINAL_SCRIPT, LEGACY_BRIDGE_SCRIPT]:
        check_file(path)


def check_legacy_workbook_state() -> None:
    if not LEGACY_WORKBOOK_FILE.exists():
        fail("Existing CrewPay_Ledger_Workbook.xlsx is missing.")


def check_protected_files_clean() -> None:
    protected_paths = [
        "CrewPay_Ledger_Workbook.xlsx",
        "apps_script/Code.gs",
        "apps_script/CrewPay_Ledger_ORIGINAL_FINAL.gs",
        "apps_script/CrewPay_Ledger_BRIDGE.gs",
    ]
    for path in protected_paths:
        check_git_path_clean(path)

    release_package = subprocess.run(
        ["git", "diff", "--quiet", "--", "release-package"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    if release_package.returncode != 0:
        fail("Protected release-package files have uncommitted changes.")

    release_untracked = subprocess.run(
        ["git", "ls-files", "--others", "--exclude-standard", "--", "release-package"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=False,
    )
    if release_untracked.stdout.strip():
        fail("Protected release-package files have untracked changes.")


def main() -> int:
    check_file(WORKBOOK_FILE)
    check_script()
    check_legacy_scripts_unchanged()
    check_legacy_workbook_state()
    check_protected_files_clean()

    wb = load_workbook(WORKBOOK_FILE, data_only=False)
    if wb.sheetnames != EXPECTED_SHEETS:
        fail(f"Workbook tab order mismatch: {wb.sheetnames!r}")

    for sheet_name, headers in EXPECTED_HEADERS.items():
        check_sheet_headers(wb[sheet_name], headers)

    check_dashboard_formulas(wb["Owner Dashboard"])
    check_data_validations(wb)
    check_forbidden_terms_in_columns(wb)
    check_forbidden_patterns(wb)

    print("CrewPay owner control workbook audit passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
